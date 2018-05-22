from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import PatternFill
import ciscoconfparse as c
import re

import sys
sys.path.insert(0, 'utils')

from get_site_data import get_site_configs, SITES_CONFIG_FOLDER

def save_wb(wb, dest_path, file_name):
    import os
    filepath = dest_path + file_name
    if not os.path.exists(dest_path):
        os.makedirs(dest_path)
    wb.save(filepath)

# +-----0-A------+-----1-B------+------2-C------+---3-D--+---4-E-+-----5-F----+-------6-G---+-------7-H---------+-------8-I-----+-------9-J-----+-------10-K-----+----11-L----+-----12-M-------+---13-N----------+---14-O----------+---15-P----------+
# +--SRC_OSW_IF--+--DST_VCE_IF--+--Access-Type--+--VLAN--+--QoS--+--Nexus_AP--+--Member/PO--+-----Descr---------+----Duplex-----+-----Speed-----+---Media Type---+---Action---+---Root-Guard---+---System-type---+---Check_Descr---+----Temp---------+
#                                   |
#                                   |
#                                   +-- Access, trunk, infra,


def atoi(text):
    ''' from string to int'''
    return int(text) if text.isdigit() else text


def natural_keys(text):
    '''
    alist.sort(key=natural_keys) sorts in human order
    http://nedbatchelder.com/blog/200712/human_sorting.html
    (See Toothy's implementation in the comments)
    '''
    return [atoi(c) for c in re.split('(\d+)', text)]


def get_col(ws, col):
    ''' Take a worksheet, return column "col" as lists '''

    return [str(ws.cell(row=r, column=col).value) for r in range(2, ws.max_row + 1)]


def get_string_from_range_to_list(range_str):
    ''' Takes '1-4' and Returns "1,2,3,4" '''

    help_list = range_str.split('-')
    start = int(help_list[0])
    stop = int(help_list[1])
    my_l = range(start, stop + 1)
    stringed_list = [str(x) for x in my_l]
    s = ','.join(stringed_list)
    return s


def get_allowed_vlan_list(if_cfg, SEL):
    ''' Get interface configuration block as a list
        and returns a list (SEL = 'LIST') or string (SEL = "STRING")
        of trunk allowed VLANS '''

    s = ''

    for line in if_cfg:

        if line[:30] == " switchport trunk allowed vlan":
            if line[:30] == " switchport trunk allowed vlan" and line[31:34] != "add":
                help_string = line[31:]
                help_string = str.rstrip(help_string)

                help_list = help_string.split(',')

                for elem in help_list:
                    if re.findall('-', elem):
                        s = s + get_string_from_range_to_list(elem) + ','
                    else:
                        s = s + elem + ','

            elif line[:34] == " switchport trunk allowed vlan add":
                help_string = line[35:]
                help_string = str.rstrip(help_string)

                help_list = help_string.split(',')

                for elem in help_list:
                    if re.findall('-', elem):
                        s = s + get_string_from_range_to_list(elem) + ','
                    else:
                        s = s + elem + ','
        else:
            continue
    s1 = s[:-1]
    if SEL == "STRING":
        return s1
    elif SEL == "LIST":
        return s1.split(',')


def get_access_vlan(if_cfg):
    ''' Get interface configuration block as a list
        and returns the access VLAN as integer '''

    for line in if_cfg:
        if line[:23] == " switchport access vlan":
            access_vlan = int(line[24:])
            break
    return access_vlan


def description_are_equals(desc_from_xls, if_cfg):
    ''' Get interface configuration block as a list + string desc_from_xls
        and returns True is are equal, False otherwise '''

    for elem in if_cfg:
        if elem[1:12] == 'description':
            desc_from_cfg = str.strip(elem[13:])
            if desc_from_xls == desc_from_cfg:
                return True
            else:
                return False


def get_channel_group(if_cfg):
    ''' Get interface configuration block as a list
        and returns the channel-group id as integer '''

    for elem in if_cfg:
        if elem[1:14] == "channel-group":
            ch_gr = int(re.findall(r'\d+', elem)[0])
    return ch_gr


def create_legendas(OUTPUT_XLS):
    wb = load_workbook(OUTPUT_XLS)
    create_qos_legendas(wb, OUTPUT_XLS)
    create_color_legendas(wb, OUTPUT_XLS)
    create_check_legendas(wb, OUTPUT_XLS)
    create_ap_legendas(wb, OUTPUT_XLS)


def create_qos_legendas(my_wb, OUTPUT_XLS):
    QOS_SHEET = 'QoS Legenda'
    ws = my_wb.create_sheet(index=1, title=QOS_SHEET)

    ws['A1'] = 'QoS Codes'
    ws['A2'] = 'U = Untrusted (set DSCP = 0 to all traffic on interface): on all ports facing OAM LAN'
    ws['A3'] = 'T = Trusted (do not change any DSCP Value): on all ports facing LTE nodes (SecGW,MME,P-GW) and IT world'
    ws['A4'] = 'V = Voice (set DSCP = EF to all traffic on interface): on all ports facing VOICE services (RTP, VOIP)'
    ws['A5'] = 'S = Signalling (set DSCP = AF31 to all traffic on interface): on all ports facing SIGNALLING services (SIP,SCTP, etc)'
    ws['A6'] = 'K = Trunk (set DSCP = AF31 for Signalling and DSCP = 0 for O&M, ACL based) on Nexus 9K'

    my_wb.save(filename=OUTPUT_XLS)


def create_color_legendas(my_wb, OUTPUT_XLS):
    COLOR_SHEET = 'Color Legenda'
    ws = my_wb.create_sheet(index=2, title=COLOR_SHEET)

    redFill = PatternFill(start_color='FF0000', end_color='FF0000',
                          fill_type='solid')      # To be Deleted
    orangeFill = PatternFill(
        start_color='FF8000', end_color='FF8000', fill_type='solid')   # To be Checked
    yellowFill = PatternFill(
        start_color='FFFF00', end_color='FFFF00', fill_type='solid')   # To be Merged
    pinkFill = PatternFill(
        start_color='eeaaee', end_color='eeaaee', fill_type='solid')     # To be Verified
    greenFill = PatternFill(start_color='a7bd2f', end_color='a7bd2f',
                            fill_type='solid')   # Not To be Verified

    ws['A1'] = 'Legend'
    ws['A2'] = 'To be Deleted??'
    ws['A2'].fill = redFill
    ws['A3'] = 'To be Checked'
    ws['A3'].fill = orangeFill
    ws['A4'] = 'To be Merged??'
    ws['A4'].fill = yellowFill
    ws['A5'] = 'Interface description'
    ws['A6'] = 'To be Verified'
    ws['A6'].fill = pinkFill
    ws['A7'] = 'Not To be Verified??'
    ws['A7'].fill = greenFill

    my_wb.save(filename=OUTPUT_XLS)


def create_check_legendas(my_wb, OUTPUT_XLS):
    CHECK_SHEET = 'Check Legenda'
    ws = my_wb.create_sheet(index=3, title=CHECK_SHEET)

    ws['A1'] = 'Checks to be done on Interfaces:'
    ws['A2'] = 'Check Half/Full duplex (Action column help here)'
    ws['A3'] = 'Change 10Mb/s --> 100Mb/s and half to full duplex where possible,'
    ws['A4'] = 'Check if notconnect ports (from show interface status command) have to migrated or not'

    my_wb.save(filename=OUTPUT_XLS)


def create_ap_legendas(my_wb, OUTPUT_XLS):
    AP_SHEET = 'Access Point Legenda'
    ws = my_wb.create_sheet(index=1, title=AP_SHEET)

    ws['A1'] = 'Access Point Values (Column C)'
    ws['A2'] = 'Access'
    ws['A3'] = 'Trunk'

    ws['C1'] = 'System Type Values (Column N)'
    ws['C2'] = 'Core-Router'
    ws['c3'] = 'Core-Switch'
    ws['C4'] = 'Decomissioned'
    ws['C5'] = 'Edge-Router'
    ws['C6'] = 'Edge-Switch'
    ws['C7'] = 'L2-Host'
    ws['C8'] = 'Monitoring'
    ws['C9'] = 'Port-Channel'
    ws['C10'] = 'Spare'

    my_wb.save(filename=OUTPUT_XLS)


def colour_output_xlsx(SHEET, OUTPUT_XLS):
    '''Get OUTPUT_XLS and  colors lines to help people on check interfaces '''

    wb = load_workbook(OUTPUT_XLS)
    ws = wb[SHEET]

    MAX_COL = ws.max_column - 1

    MAX_COLUMN_COLOR = MAX_COL

    redFill = PatternFill(start_color='FF0000', end_color='FF0000',
                          fill_type='solid')         # To be Deleted
    orangeFill = PatternFill(
        start_color='FF8000', end_color='FF8000', fill_type='solid')   # To be Checked
    yellowFill = PatternFill(
        start_color='FFFF00', end_color='FFFF00', fill_type='solid')   # To be Merged
    pinkFill = PatternFill(start_color='eeaaee',
                           end_color='eeaaee', fill_type='solid')

    for row in ws.rows:
        #src_if = row[0].value

        if row[14].value == "Description unchanged":    # same description

            # if ((portchannel or member) and (id in INFRA)list)) or marked as
            # infra then red
            if (row[5].value == "Infra" and str.isdigit(str(row[6].value))) or row[13].value == "Decommissioned" or row[13].value == "Spare" or row[13].value == "Monitoring" or row[3].value == 1:
                for cell in row[0:MAX_COLUMN_COLOR]:
                    cell.fill = redFill
            elif (row[5].value == "Infra" and not(str.isdigit(str(row[6].value)))) or row[13].value == "TBV" or row[13].value == "TBV-NC":
                for cell in row[0:MAX_COLUMN_COLOR]:
                    cell.fill = orangeFill
            elif row[13].value == "Core-Router" or row[13].value == "Core-Switch":
                for cell in row[0:MAX_COLUMN_COLOR]:
                    cell.fill = yellowFill
        else:                                           # different description
            for cell in row[0:MAX_COLUMN_COLOR]:
                cell.fill = pinkFill

    wb.save(filename=OUTPUT_XLS)
    print("End F2")


def get_descr(if_cfg):

    desc_from_cfg = ''
    for elem in if_cfg:
        if elem[1:12] == 'description':
            desc_from_cfg = str.strip(elem[13:])
    return desc_from_cfg


def readin_xls_writeout_xls(OSW_CFG_TXT, INPUT_XLS, SHEET, OUTPUT_XLS, box_config):

    header_out = ['SRC OSW IF', 'DST VCE IF', 'Access Type', 'VLAN', 'QoS', 'Nexus AP', 'Member/PO',
                  'Descr', 'Duplex', 'Speed', 'Media Type', 'Action', 'Root-Guard', 'System Type', 'Check Descr']

    parse = c.CiscoConfParse(OSW_CFG_TXT)

    intf_obj_list = parse.find_objects(r'^interface')
    wb_r = load_workbook(INPUT_XLS)
    wb_w = Workbook()

    ws_r = wb_r[SHEET]
    ws_w = wb_w.create_sheet(index = 0, title = SHEET)
    
    MAX_COL = 15
    MAX_ROW = ws_r.max_row

    pinkFill = PatternFill(
        start_color='eeaaee', end_color='eeaaee', fill_type='solid')     # To be Verified
    greenFill = PatternFill(start_color='a7bd2f',
                            end_color='a7bd2f', fill_type='solid')

    ws_w.append(header_out)

    # WB_W Table Exist if it has been accessed
    ws_w.cell(row=MAX_ROW, column=MAX_COL, value=10).value

    for row_r, row_w in zip(ws_r.rows, ws_w.rows):
        if row_r[0].value == "Device":
            continue
        # was intf_from_xls = 'interface ' + str.strip(str(row_r[3].value))
        intf_from_xls = str.strip(str(row_r[3].value))
        #print row_r[3].value
        # Copy interface (or row_r[3].value)
        row_w[0].value = str(row_r[3].value)
        # Copy New-Nexus AP into Nexus AP
        row_w[5].value = str(row_r[13].value)
        # row_w[7].value = row_r[5].value
        # # Copy Port Description
        # Copy Duplex
        row_w[8].value = str(row_r[8].value)
        # Copy Speed
        row_w[9].value = str(row_r[9].value)
        # Copy Port Media_type
        row_w[10].value = str(row_r[10].value)
        # Copy Action
        row_w[11].value = str(row_r[17].value)

        # copy System Type
        row_w[13].value = str(row_r[12].value)

        for intf_obj in intf_obj_list:
            # IOS IF's Configuration
            intf_cfg = intf_obj.ioscfg
            # First line is IF itself
            intf_from_cfg = str.strip(intf_cfg[0])

            if intf_from_xls == intf_from_cfg:

                if intf_from_xls[0:22] == "interface Port-channel":
                    po = int(re.findall(r'\d+', intf_from_xls)[0])
                    row_w[6].value = po

                if intf_obj.has_child_with("switchport trunk allowed vlan"):
                    vlan_list_string = get_allowed_vlan_list(
                        intf_cfg, "STRING")
                    row_w[2].value = 'Trunk'
                    row_w[3].value = vlan_list_string

                    if intf_obj.has_child_with("channel-group"):
                        channel_group = get_channel_group(intf_cfg)
                        row_w[6].value = channel_group

                elif intf_obj.has_child_with("switchport mode access") or intf_obj.has_child_with("switchport access vlan"):
                    row_w[2].value = 'Access'
                    if intf_obj.has_child_with("switchport access vlan"):
                        access_vlan = get_access_vlan(intf_cfg)
                        row_w[3].value = str(access_vlan)
                    else:
                        row_w[3].value = '1'
                elif not(intf_obj.has_child_with("switchport mode")) and intf_obj.has_child_with("shutdown"):
                    row_w[2].value = 'ShutDown'
                    row_w[3].value = '1'

                if intf_obj.has_child_with("description"):
                    if description_are_equals(str.strip(str(row_r[5].value)), intf_cfg):
                        row_w[7].value = str(row_r[5].value)
                        row_w[14].value = "Description unchanged"
                        row_w[14].fill = greenFill
                    else:
                        # row_w[7].value = "INTERFACE TO BE CHECKED"
                        row_w[7].value = get_descr(intf_cfg)
                        row_w[14].value = "Description CHANGED!!!"
                        row_w[14].fill = pinkFill
                elif not(intf_obj.has_child_with("description")):
                    if row_r[5].value is None:
                        row_w[14].value = "Description unchanged"
                        row_w[14].fill = greenFill
                    else:
                        # row_w[7].value = "INTERFACE TO BE CHECKED"
                        row_w[7].value = get_descr(intf_cfg)
                        row_w[14].value = "Description CHANGED!!!"
                        row_w[14].fill = pinkFill
            else:
                continue

        # Copy Root_Guard
        import json
        vlan_without_spt = []
        with open(box_config.base_dir + box_config.site + "/DATA_SRC/CMD/" +
                  box_config.switch + '_vlan_similar_to_4093.txt') as f:
            vlan_without_spt = json.load(f)

        # Copy vlan lower than 3900
        for vlan_no_spt in vlan_without_spt:
            vlans_in_excel = str(row_w[3].value).split(',')
            if vlan_no_spt in vlans_in_excel:
                for vlan in vlans_in_excel:
                    if int(vlan) < 3900:
                        VlanProblem(row_w[0].value, box_config.switch)
                row_w[12].value = "No"
            else:
                row_w[12].value = str(row_r[11].value)
    wb_w.save(filename=OUTPUT_XLS)
    print("End F1")

def VlanProblem(interface, box):
    try:
        raise ValueError('Vlan lower than 3900 is on the same trunk '
                         'of a Vlan similar to 4000 on: ' + interface + ' on ' + box)
        raise Exception('VlanProblem')
    except Exception as error:
        raise

def further_interfaces(site_config, OSW_CFG_TXT, SHEET, OUTPUT_XLS):
    parse = c.CiscoConfParse(OSW_CFG_TXT)

    intf_obj_list = parse.find_objects(r'^interface .*Ethernet|^interface Port-channel.*')
    if_list_cfg = [intf.text for intf in intf_obj_list if len(intf.ioscfg) > 3]

    wb = load_workbook(OUTPUT_XLS)

    ws = wb[SHEET]
    MAX_ROW = ws.max_row
    START_ROW = MAX_ROW + 1

    if_list_xls = get_col(ws, 1)

    len_if_list_xls = len(if_list_xls)
    len_if_list_cfg = len(if_list_cfg)

    if_set_xls = set(if_list_xls)
    if_set_cfg = set(if_list_cfg)

    if (len_if_list_cfg - len_if_list_xls) > 0:

        cfg_less_xls_list = list(if_set_cfg - if_set_xls)
        cfg_less_xls_list.sort(key=natural_keys)

        ws.cell(row=START_ROW, column=1).value = "Le seguenti interfaccie sono utilizzate nella cfg ma non compaiono nell' xlsx"

        for elem, line in zip(cfg_less_xls_list, range(START_ROW + 1, START_ROW + 1 + len(cfg_less_xls_list))):
            ws.cell(row=line, column=1).value = elem
    elif (len_if_list_cfg - len_if_list_xls) == 0:
        ws.cell(row=START_ROW, column=1).value = "Non Vi sono interfaccie aggiuntive"
    elif (len_if_list_cfg - len_if_list_xls) < 0:

        xls_less_cfg_list = list(if_set_xls - if_set_cfg)
        xls_less_cfg_list.sort(key=natural_keys)
        ws.cell(row=START_ROW, column=1).value = "Le seguenti interfaccie sono presenti nell XLSX ma non nella cfg--> VERIFICARE"

        for col in range(1, 16):
            ws.cell(row=START_ROW, column=col).fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC',
                                                                  fill_type='solid')

        for elem, line in zip(xls_less_cfg_list, range(START_ROW + 1, START_ROW + 1 + len(xls_less_cfg_list))):
            ws.cell(row=line, column=1).value = elem

    wb.save(OUTPUT_XLS)
    dest_path = site_config.base_dir + site_config.site + "/DATA_SRC/XLS/OUTPUT_STAGE_1/"
    save_wb(wb, dest_path, site_config.switch + '_OUT_DB.xlsx')

def run(site_configs):
    for box_config in site_configs:

        base_dir = box_config.base_dir + box_config.site + box_config.switch + "/Stage_1/"
        INPUT_XLS = base_dir + box_config. switch + '_DB_MIGRATION.xlsx'
        OUTPUT_XLS = base_dir + box_config.switch + '_OUT_DB.xlsx'
        OSW_CFG_TXT = base_dir + box_config.switch + '.txt'

        SHEET = box_config.sheet

        readin_xls_writeout_xls(OSW_CFG_TXT, INPUT_XLS, box_config.sheet, OUTPUT_XLS, box_config)
        colour_output_xlsx(box_config.sheet, OUTPUT_XLS)
        further_interfaces(box_config, OSW_CFG_TXT, box_config.sheet, OUTPUT_XLS)
        create_legendas(OUTPUT_XLS)
        print('End script')

#copies config and excel files in Stage_1 folder from DATA_SRC folder
def copy_files(box_configs):
    import shutil
    import os
    for box_config in box_configs:

        # copies config files  in Stage1 folder
        source = box_config.conf_dest_path[1] + box_config.switch + ".txt"
        dest = box_config.conf_dest_path[0] + box_config.switch + ".txt"
        if not os.path.exists(box_config.conf_dest_path[0]):
            os.makedirs(box_config.conf_dest_path[0])
        shutil.copy(source, dest)
        print(box_config.switch + ".txt copied in " + box_config.conf_dest_path[1] +
              box_config.switch + '.txt' + ".")

        # copies excel files  in Stage1 folder
        source = box_config.new_excel_file_paths[1] + box_config.switch + "_DB_MIGRATION.xlsx"
        dest = box_config.new_excel_file_paths[0] + box_config.switch + "_DB_MIGRATION.xlsx"
        shutil.copy(source, dest)
        print(box_config.switch + "_DB_MIGRATION.xlsx copied in " + box_config.new_excel_file_paths[0] +
              box_config.switch + "_DB_MIGRATION.xlsx" ".")


if __name__ == "__main__":
    site_configs = get_site_configs(SITES_CONFIG_FOLDER)
    copy_files(site_configs)
    run(site_configs)