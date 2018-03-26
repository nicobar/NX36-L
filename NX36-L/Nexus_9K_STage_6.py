import openpyxl
import sys
from copy import copy

sys.path.insert(0, 'utils')

VLAN_NUMBER = 4096
NUMBER_OF_INTTERFACES = 300

from get_site_data import get_site_configs, SITES_CONFIG_FOLDER, exists

def get_excel_sheet(filename, sheet_name):
    wb = openpyxl.load_workbook(filename)
    return wb[sheet_name]

def copy_sheet(new, template, start_col, start_row, max_row):
    for row in range(start_row, max_row):
        for column in range(ord(start_col) - ord('A'), ord('Z') - ord('A')):
            col = chr(ord('A') + column)
            cell = "{}{}".format(col, row)
            new[cell].value = copy(template[cell].internal_value)
            new[cell].font = copy(template[cell].font)
            new[cell].border = copy(template[cell].border)
            new[cell].fill = copy(template[cell].fill)
            new[cell].number_format = copy(template[cell].number_format)
            new[cell].protection = copy(template[cell].protection)
            new[cell].alignment = copy(template[cell].alignment)

def parse_vrrp(file_cmd):
    '''
    Get all vrrp vlan from file_cmd XXX_show_vrrp_brief
    :param file_cmd:
    :return:
    '''
    import re
    with open(file_cmd, encoding="utf-8") as file:
        config = file.read()
        config = config.split("\n")
        vlans = {}
        for line in config:
            import os
            line = os.linesep.join([s for s in line.splitlines() if s])
            if line.startswith('Vl'):
                vlans_id = re.search(r'Vl(\d+)', line).group(1)
                line = line.split("  ")
                vlans[vlans_id] = [line[-1]]
    return vlans


def create_vlan_voice_line(site_configs, ws):
    import json

    output = {}
    for site_config in site_configs:
        file_cmd = site_config.base_dir + site_config.site + 'DATA_SRC/CMD/' + \
                   site_config.switch + '_voice_vlan_data.txt'
        with open(file_cmd, encoding="utf-8") as file:
            voice_vlans = json.load(file)
            for voice_vlan in voice_vlans:
                if voice_vlan not in output:
                    output[voice_vlan] = voice_vlans[voice_vlan]
                else:
                    for k in range(1, len(voice_vlans[voice_vlan])):
                        output[voice_vlan][k] = output[voice_vlan][k] + " " + voice_vlans[voice_vlan][k]
    row = 5
    cols = ['E','F','G', 'H', 'I']
    for vlan in output:
        k = 0
        cell = "{}{}".format(cols[k], row)
        ws[cell] = vlan
        for data in output[vlan]:
            k = k + 1
            cell = "{}{}".format(cols[k], row)
            ws[cell].value = data
        row = row + 1

def check_vlan(vrrp_vlans, vce):
    '''
    checks if vrrp vlan has a static route. if so it adds it to the
    vrrp_vlans dictionary with a list that consists of ip mask netx and mask
    :param vrrp_vlans:
    :param vce:
    :return:
    '''
    import re
    with open(vce, encoding="utf-8") as file:
        config = file.read()
        static_pos = config.find('####')
        static_block = config[static_pos:]
        static_block = static_block.split("\n")

        for line in static_block[2:]:
            import os
            line = os.linesep.join([s for s in line.splitlines() if s])
            if 'ip route' in line:
                data = re.search(r'ip route (\d+\.\d+\.\d+\.\d+) (\d+\.\d+\.\d+\.\d+) Vlan(\d+)', line)
                vlan = data.group(3)

                if vlan in vrrp_vlans:
                    ip = data.group(1)
                    mask = data.group(2)
                    bare_network = net_calc(ip, mask)

                    vrrp_vlans[vlan].append(bare_network)
                    vrrp_vlans[vlan].append(mask)
                    vrrp_vlans[vlan].append('static')
    return vrrp_vlans

def net_calc(ip, mask):
    import ipcalc

    addr = ipcalc.IP(ip, mask)
    network_with_cidr = str(addr.guess_network())
    bare_network = network_with_cidr.split('/')[0]
    #print(addr, network_with_cidr, bare_network)

    return bare_network


def get_subnet_for_non_static(vrrp_vlans, osw):
    '''
    checks the remaining direct connected netx in the osw config
    and adds it to the vrrp_vlans dictionary with a list that consists of ip mask netx and mask
    :param vrrp_vlans:
    :param osw:
    :return:
    '''
    import re
    with open(osw, encoding="utf-8") as file:
        config = file.read()
        config = config.split("!")

        remaining_vlans = []
        for vlan in vrrp_vlans:
            if len(vrrp_vlans[vlan]) == 1:
                remaining_vlans.append(vlan)

        for block in config:
            import os
            block = os.linesep.join([s for s in block.splitlines() if s])
            if block.startswith('interface Vlan'):
                vlans_id = re.search(r'Vlan(\d+)', block).group(1)
                if vlans_id in remaining_vlans:
                    block = block.split("\n")
                    for line in block:
                        if 'ip address' in line:
                            data = re.search(r'ip address (\d+\.\d+\.\d+\.\d+) (\d+\.\d+\.\d+\.\d+)', line)
                            ip = data.group(1)
                            mask = data.group(2)
                            bare_network = net_calc(ip, mask)

                            vrrp_vlans[vlans_id].append(bare_network)
                            vrrp_vlans[vlans_id].append(mask)
                            vrrp_vlans[vlans_id].append('connected')
    return vrrp_vlans

def fill_static_route(site_configs, ws):
    '''
        fill static route
    '''
    list_of_tuple = []

    for site_config in site_configs:
        vce = site_config.base_dir + site_config.site + 'AID/' + \
                       site_config.switch + 'VCE.txt'
        with open(vce, encoding="utf-8") as file:
            config = file.read()
            static_pos = config.find('####')
            static_block = config[static_pos:]
            static_block = static_block.split("\n")

            for line in static_block[2:]:
                import os
                line = os.linesep.join([s for s in line.splitlines() if s])
                if 'ip route' in line:
                    list_of_tuple.append((site_config.switch, line))

    row = 5
    col = ['F','G']
    for line in list_of_tuple:
        cell = "{}{}".format('F', row)
        ws[cell] = line[0]
        cell = "{}{}".format('G', row)
        ws[cell] = line[1]
        row = row + 1



def remove_static_from_final(site_configs):
    '''
        fill static route
    '''
    list_of_tuple = []

    for site_config in site_configs:
        vce = site_config.base_dir + site_config.site + 'FINAL/' + \
                       site_config.switch + 'VCE.txt'
        clean_config = ''
        with open(vce, encoding="utf-8") as file:
            config = file.read()
            static_pos = config.find('####')
            clean_config = config[0:static_pos]


        with open(vce, 'w') as f:
           f.write(clean_config)
           f.close()


def fill_vrrp_excel(final, ws):
   row = 8
   cols = ['F','G','I','J','K']
   for vlan in final:
       k = 0
       cell = "{}{}".format(cols[k], row)
       ws[cell] = vlan
       for data in final[vlan]:
           k = k + 1
           cell = "{}{}".format(cols[k], row)
           ws[cell].value = data
       row = row + 1

def remove_not_migrated_vlan(vrrp_vlans, vce):
    '''
       removes not migrated vlan
    '''
    import re
    with open(vce, encoding="utf-8") as file:
        config = file.read()
        config = config.split("!")

        cleand_vrrp_vlan = {}
        for block in config:
            import os
            block = os.linesep.join([s for s in block.splitlines() if s])
            if block.startswith('vlan '):
                vlans_id = re.search(r'vlan (\d+)', block).group(1)
                if vlans_id in vrrp_vlans:
                    cleand_vrrp_vlan[vlans_id] = vrrp_vlans[vlans_id]
    return cleand_vrrp_vlan

def run(site_configs):
    import re

    AID_PATH = site_configs[0].base_dir + site_configs[0].site + 'AID/'
    FINAL_PATH = site_configs[0].base_dir + site_configs[0].site + 'FINAL/'

    EXCEL_SRC = site_configs[0].base_dir + site_configs[0].site + 'DATA_SRC/XLS/OUTPUT_STAGE_2.0/'
    CMD_PATH = site_configs[0].base_dir + site_configs[0].site + 'DATA_SRC/CMD/' + \
               site_configs[0].switch + '_show_vrrp_brief.txt'

    CFG_PATH_VCE = site_configs[0].base_dir + site_configs[0].site + 'AID/' + \
                   site_configs[0].switch + 'VCE.txt'

    CFG_PATH_OSW = site_configs[0].base_dir + site_configs[0].site + 'DATA_SRC/CFG/' + \
                   site_configs[0].switch + '.txt'

    site = site_configs[0].switch
    m = re.search('([A-Z]{2})OSW(\d\d)', site)
    site = m.group(1) + m.group(2)

    INPUT_XLS = AID_PATH + 'AID_to_{}_NMP.xlsx'.format(site)
    OUTPUT_XLS = FINAL_PATH + '{}_MIGRAZIONE_TABLES.xlsx'.format(site)

    migration_table = openpyxl.Workbook()


    ####### STATIC ROUTE SHEET ##################
    sheet = 'Static Routes (OSW)'
    ws = migration_table.create_sheet(sheet)

    fill_static_route(site_configs, ws)
    remove_static_from_final(site_configs)

    #########  VRRP SHEET  #######################
    sheet = '{} VRRP VIP & STATIC'.format(site)
    ws = migration_table.create_sheet(sheet)
    cell = "{}{}".format("F", "6")
    ws[cell] = site_configs[0].switch + '& ' + site_configs[1].switch
    cell = "{}{}".format("I", "6")
    ws[cell] = 'Statiche traffico intra-sito'
    cell = "{}{}".format("F", "7")
    ws[cell] = 'VLAN'
    cell = "{}{}".format("G", "7")
    ws[cell] = 'VRRP IP'
    cell = "{}{}".format("I", "7")
    ws[cell] = 'Net X'
    cell = "{}{}".format("J", "7")
    ws[cell] = 'MASK'
    cell = "{}{}".format("K", "7")
    ws[cell] = 'Note'

    vlans = parse_vrrp(CMD_PATH)
    if len(vlans) > 0:
        vrrp_vlans = remove_not_migrated_vlan(vlans, CFG_PATH_VCE)
        vrrp_vlans = check_vlan(vrrp_vlans, CFG_PATH_VCE)
        final = get_subnet_for_non_static(vrrp_vlans, CFG_PATH_OSW)
        fill_vrrp_excel(final, ws)


    ####### VOICE VLAN ROUTE SHEET ##################
    sheet = '{} Voice VLAN'.format(site)
    ws = migration_table.create_sheet(sheet)
    create_vlan_voice_line(site_configs, ws)

    #########COPY SHEET INTERFACE #######################
    for site_config in site_configs:
        INPUT2_XLS = EXCEL_SRC + '{}_OUT_DB_OPT.xlsx'.format(site_config.switch)
        int_xls = openpyxl.load_workbook(INPUT2_XLS)
        int_sheet = int_xls
        sheet = '{}_IF'.format(site_config.switch)
        int_sheet = int_xls[site_config.switch]
        ws = migration_table.create_sheet(sheet)
        copy_sheet(ws, int_sheet, 'A', 1, NUMBER_OF_INTTERFACES)

    #########COPIA SHEET AID - DOES NOT WORK #######################

    matrix_sheet = 'VLAN_Migration_Matrix'
    stp_sheet = 'Higher STP Cost IF'

    aid_file_xls = openpyxl.load_workbook(INPUT_XLS)
    STP_Cost = aid_file_xls[stp_sheet]
    VLAN_Migration_Matrix = aid_file_xls[matrix_sheet]

    ws = migration_table.create_sheet(matrix_sheet)
    # this copy does not work
    # copy_sheet(ws, VLAN_Migration_Matrix, 'E', 4, VLAN_NUMBER)
    ws = migration_table.create_sheet(stp_sheet)
    # copy_sheet(ws, STP_Cost, 'D', 6, VLAN_NUMBER)

    migration_table.save(filename=OUTPUT_XLS)

if __name__ == "__main__":
    site_configs = get_site_configs(SITES_CONFIG_FOLDER)
    run(site_configs)