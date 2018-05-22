from openpyxl import load_workbook
import ciscoconfparse as ccp
import pprint

class VlanListCondtioned:
    
    def __init__(self, xls_file_pathname, po_osw_osw,po_list, sheet, interest_column, condition_column, condition):
        self.xls_file_pathname = xls_file_pathname
        self.po_osw_osw = po_osw_osw
        self.sheet = sheet
        self.interest_column = interest_column
        self.condition_column = condition_column
        self.condition = condition
        self.vlan_accestrunk_list = self.get_column_conditioned()
        #self.po = po
        self.po_list = po_list
        self.po_vlan_allowed_db = self.get_po_allowed_vlan_db()
        self.new_vlan_allowed_db = self.elaborate_new_po_allowed_list()
        
    def __from_range_to_list(self, range_str):
        ''' from '1-4' to '1,2,3,4' '''
        mylist = []
    
        h_l = range_str.split('-')
        start = int(h_l[0])
        stop = int(h_l[1])
        for x in range(start, stop + 1):
            mylist.append(str(x))
        return mylist
    
    def __order_list_of_numb_str(self, lst):
        mystr = [int(elem) for elem in lst]
        mystr.sort()
        return mystr
    
    def __list_of_numb_to_str(self, lst):
        mystr = [str(elem) for elem in lst]
        mystr = ','.join(mystr)
        return mystr

    def get_column_conditioned(self):
        ''' Take a worksheet, return column "interest_column" as lists conditioned to "condition" of condition_column'''
        
        vlan_list = []
        
        wb = load_workbook(self.xls_file_pathname)
        ws = wb.get_sheet_by_name(self.sheet)
        
        mylist = [str(ws.cell(row=r, column=self.interest_column).value) for r in range(3, ws.max_row) if ws.cell(row=r, column=self.condition_column).value == int(self.condition)]
        
        wb.close()
        
        for elem in mylist:
            if ',' in elem:
                token = elem.split(',')
                for tk in token:
                    if '-' in tk:
                        ml = self.__from_range_to_list(tk)
                        for el in ml:
                            vlan_list.append(self.__from_range_to_list(el))
                    else:
                        vlan_list.append(tk)
                        
            elif '-' in elem:
                ml = self.__f_rom_range_to_list(elem)
                for el in ml:
                    vlan_list.append(self.__from_range_to_list(el))
        
            else:
                vlan_list.append(elem)
    
        vlan_set = set(vlan_list)
        vlan_list = list(vlan_set)
        vlan_list = self.__order_list_of_numb_str(vlan_list)
        return vlan_list
    
    def get_common_vlan(self, other):
        return [elem for elem in self.vlan_accestrunk_list if elem in other.vlan_accestrunk_list]
        
    
    def get_po_allowed_vlan_db(self):
        
        wb = load_workbook(self.xls_file_pathname)
        ws = wb.get_sheet_by_name(self.sheet)
        
        # {'po1': allowed_vlan_list_po1, etc}
        po_db = {}
        
        
        for row in ws.iter_rows(row_offset = 1):
            if row[0].value is None:
                break
            else:
                _,intf = str(row[0].value).split(' ')
                for po in self.po_list:
                    if  intf == po:
                        po_db[intf]  = row[3].value
        return po_db
    
    def elaborate_new_po_allowed_list(self):
        
        new_po_db = {}
    
        for po, lst_s in self.po_vlan_allowed_db.items():
            lst = lst_s.split(',') 
            lst = [int(elem) for elem in lst]
            new_allowed_lst = [str(elem) for elem in lst if elem in self.vlan_accestrunk_list]
            new_allowed_str = ','.join(new_allowed_lst)
            new_po_db[po] = new_allowed_str
        return new_po_db
    
    def get_possible_mono_vlan_db(self, other):
        po1_vlan_OSW1 = self.new_vlan_allowed_db[self.po_osw_osw]
        po1_vlan_OSW2 = other.new_vlan_allowed_db[self.po_osw_osw]
        a = set(po1_vlan_OSW1.split(',')) 
        b = set(po1_vlan_OSW2.split(','))
        possible_mono_vlan = a ^ b
        return possible_mono_vlan

    def get_po_list_union(self, other):
        po1_vlan_OSW1 = self.new_vlan_allowed_db[self.po_osw_osw]
        po1_vlan_OSW2 = other.new_vlan_allowed_db[self.po_osw_osw]
        a = set(po1_vlan_OSW1.split(',')) 
        b = set(po1_vlan_OSW2.split(','))
        po_vce_vce_vlan = a | b
        po_vce_vce_vlan = self.__order_list_of_numb_str(list(po_vce_vce_vlan))
        po_vce_vce_vlan  = self.__list_of_numb_to_str(po_vce_vce_vlan)
        return po_vce_vce_vlan
        

file_path = '/mnt/hgfs/VM_shared/VF-2017/NMP/Sites/RM02_preparation/COMMON/DATA_SRC/'
xls_filename_rm023  =  'XLS/OUTPUT_STAGE_1.5_RM2_and_BIS/RMOSW023_checked_v1.1_OUT_DB_OPT.xlsx'
xls_file_pathname_rm023 = file_path + xls_filename_rm023

sheet_rm023 = 'RMOSW023'
po_db_rm023 = ['Port-channel123', 'Port-channel1', 'GigabitEthernet4/15', 'GigabitEthernet4/21', 'GigabitEthernet4/22']
po_osw_osw = 'Port-channel1'

xls_filename_rm026  =  'XLS/OUTPUT_STAGE_1.5_RM2_and_BIS/RMOSW026_checked_v1.1_OUT_DB_OPT.xlsx'
xls_file_pathname_rm026 = file_path + xls_filename_rm026

sheet_rm026 = 'RMOSW026'
po_db_rm026 = ['Port-channel126', 'Port-channel1', 'GigabitEthernet4/19', 'GigabitEthernet4/22', 'GigabitEthernet4/24']


interest_column = 4     # VLAN
condition_column = 16   # DEST_VCE_COUPLE
condition_1 = 1
condition_2 = 2


RM023_cond1 = VlanListCondtioned(xls_file_pathname_rm023, po_osw_osw,po_db_rm023, sheet_rm023, interest_column, condition_column, condition_1)
RM023_cond2 = VlanListCondtioned(xls_file_pathname_rm023, po_osw_osw, po_db_rm023, sheet_rm023, interest_column, condition_column, condition_2)
RM026_cond1 = VlanListCondtioned(xls_file_pathname_rm026, po_osw_osw, po_db_rm026, sheet_rm026, interest_column, condition_column, condition_1)
RM026_cond2 = VlanListCondtioned(xls_file_pathname_rm026, po_osw_osw, po_db_rm026, sheet_rm026, interest_column, condition_column, condition_2)

print('Lista VLANs in accesso (access & trunk) per {osw} verso la coppia VCE {cp}: {lst}'.format(lst = RM023_cond1.vlan_accestrunk_list, cp = RM023_cond1.condition, osw = RM023_cond1.sheet))
print('Lista VLANs in accesso (access & trunk) per {osw} verso la coppia VCE {cp}: {lst}'.format(lst = RM026_cond1.vlan_accestrunk_list, cp = RM026_cond1.condition, osw = RM026_cond1.sheet))
print()
print('Lista VLANs in accesso (access & trunk) per {osw} verso la coppia VCE {cp}: {lst}'.format(lst = RM023_cond2.vlan_accestrunk_list, cp = RM023_cond2.condition, osw = RM023_cond2.sheet))

print('Lista VLANs in accesso (access & trunk) per {osw} verso la coppia VCE {cp}: {lst}'.format(lst = RM026_cond2.vlan_accestrunk_list, cp = RM026_cond2.condition, osw = RM026_cond2.sheet))


print ("i VCE1 della prima e seconda coppia hanno in comune le VLAN: ", RM023_cond2.get_common_vlan(RM023_cond1))
print ("i VCE1 della prima e seconda coppia hanno in comune le VLAN: ", RM026_cond2.get_common_vlan(RM026_cond1))

print()
print ("Le liste di VLAN sui vecchi trunk sono: {alwd_lst}".format(alwd_lst = RM023_cond1.po_vlan_allowed_db))
print ("Le liste di VLAN sui vecchi trunk sono: {alwd_lst}".format( alwd_lst = RM026_cond1.po_vlan_allowed_db))
print ("le lista di VLAN sul nuovo PO originato da {osw} tra l'omologo della coppia VCE {cp} e il VPE sono: {new_alwd_lst}".format( cp = RM023_cond1.condition, osw = RM023_cond1.sheet, new_alwd_lst = RM023_cond1.new_vlan_allowed_db))
print ("le lista di VLAN sul nuovo PO originato da {osw} tra l'omologo della coppia VCE {cp} e il VPE sono: {new_alwd_lst}".format( cp = RM026_cond1.condition, osw = RM026_cond1.sheet, new_alwd_lst = RM026_cond1.new_vlan_allowed_db))

print()
print ("Le liste di VLAN sui vecchi trunk sono: {alwd_lst}".format( alwd_lst = RM023_cond2.po_vlan_allowed_db))
print ("Le liste di VLAN sui vecchi trunk sono: {alwd_lst}".format( alwd_lst = RM026_cond2.po_vlan_allowed_db))
print ("le lista di VLAN sul nuovo PO originato da {osw} tra l'omologo della coppia VCE {cp} e il VPE sono: {new_alwd_lst}".format( cp = RM023_cond2.condition, osw = RM023_cond2.sheet, new_alwd_lst = RM023_cond2.new_vlan_allowed_db))
print ("le lista di VLAN sul nuovo PO originato da {osw} tra l'omologo della coppia VCE {cp} e il VPE sono: {new_alwd_lst}".format( cp = RM026_cond2.condition, osw = RM026_cond2.sheet, new_alwd_lst = RM026_cond2.new_vlan_allowed_db))

print ()

print ("Da verificare se le seguenti VLAN {mono} sono monoattestate per la coppia {cp}".format(mono = RM023_cond1.get_possible_mono_vlan_db(RM026_cond1), cp = RM023_cond1.condition))
print ("Da verificare se le seguenti VLAN {mono} sono monoattestate per la coppia {cp}".format(mono = RM023_cond2.get_possible_mono_vlan_db(RM026_cond2), cp = RM023_cond2.condition))


print ("Se sono monoattestate per la coppia {cp} la lista delle VLAN VCE-VCE è: {un}".format(un =  RM023_cond1.get_po_list_union(RM026_cond1), cp = RM023_cond1.condition))
print ("Se sono monoattestate per la coppia {cp} la lista delle VLAN VCE-VCE è: {un}".format(un =  RM023_cond2.get_po_list_union(RM026_cond2), cp = RM023_cond2.condition))

print()
print('Prova pretty print')
pprint.pprint(RM023_cond1.po_vlan_allowed_db, indent=1)