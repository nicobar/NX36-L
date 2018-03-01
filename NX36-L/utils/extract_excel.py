import json
from copy import copy
import openpyxl
from get_site_data import get_site_configs, exists, SITES_CONFIG_FOLDER

def get_excel_sheet(filename):
    wb = openpyxl.load_workbook(filename)
    first_sheet = wb.sheetnames[0]
    return wb["Summary_19JAN17"]

def create_new_excel(sheet_name):
    wb = openpyxl.Workbook()
    wb.create_sheet(sheet_name)
    return wb[sheet_name], wb

def save_wb(wb, dest_path, file_name):
    import os
    filepath = dest_path + file_name
    if not os.path.exists(dest_path):
        os.makedirs(dest_path)
    wb.save(filepath)

class Create_Excel():
    def __init__(self, excel_files, site_config):
        self.site_config = site_config
        self.original_excel = excel_files[0]
        self.new_excel = excel_files[1]
        self.switch_column = 'B'  # column corresponding to switch in the origianl excel file

    def copy_row(self, row1, row2):
        for column in range(0, ord('U') - ord('B') + 1):
            col1 = chr(ord('A') + column)
            col2 = chr(ord('B') + column)
            cell2 = "{}{}".format(col2, row2)
            cell1 = "{}{}".format(col1, row1)
            self.new_excel[cell1].value = self.original_excel[cell2].value
            if self.original_excel[cell2].has_style:
                self.new_excel[cell1].font = copy(self.original_excel[cell2].font)
                self.new_excel[cell1].border = copy(self.original_excel[cell2].border)
                self.new_excel[cell1].fill = copy(self.original_excel[cell2].fill)
                self.new_excel[cell1].number_format = copy(self.original_excel[cell2].number_format)
                self.new_excel[cell1].protection = copy(self.original_excel[cell2].protection)
                self.new_excel[cell1].alignment = copy(self.original_excel[cell2].alignment)

    def extract_info(self):
        self.copy_row(1,1)

        i = 2
        for row in range(2, self.original_excel.max_row):
            cell = "{}{}".format(self.switch_column, row)
            if self.original_excel[cell].value == self.site_config.switch:
                self.copy_row(i,row)
                i = i + 1

def get_excel(site_configs):
    import shutil

    original = None
    for box_config in site_configs:

        #it only checks if file exists in DATA_SRC folder
        if not exists(box_config.new_excel_file_paths[1] + box_config.switch + "_DB_MIGRATION.xlsx"):

            print("Excel file is about to be extracted in " + box_config.new_excel_file_paths[1] +
                  box_config.switch + "_DB_MIGRATION.xlsx" ".")

            if original == None:
                original = get_excel_sheet(box_config.base_dir + "/Migrazione/Nexus_9k_new_v0.6.xlsx")
            new_excel, wb = create_new_excel(box_config.switch)
            new_site_db = Create_Excel([original, new_excel], box_config)
            new_site_db.extract_info()

            for box in box_config.new_excel_file_paths:
                save_wb(wb, box, box_config.switch + "_DB_MIGRATION.xlsx")

            print("Excel file has been extracted for " + box_config.switch + ".")
        else:
            print("Excel file is already in place in " + box_config.new_excel_file_paths[1] +
                  box_config.switch + "_DB_MIGRATION.xlsx" ".")
            # copies the file in Stage1 folder
            source = box_config.new_excel_file_paths[1] + box_config.switch + "_DB_MIGRATION.xlsx"
            dest = box_config.new_excel_file_paths[0] + box_config.switch + "_DB_MIGRATION.xlsx"
            shutil.copy(source, dest)
            print("-> Copied in " + box_config.new_excel_file_paths[0] +
                  box_config.switch + "_DB_MIGRATION.xlsx" ".")

#if __name__ == "__main__":
#    site_configs = get_site_configs(SITES_CONFIG_FOLDER)
#    run_extract_excel(site_configs)