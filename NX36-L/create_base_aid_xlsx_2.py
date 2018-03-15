##############################################
################# IMPORTS ####################
##############################################


import ciscoconfparse as c
from openpyxl.workbook import Workbook

import sys
sys.path.insert(0, 'utils')

from get_site_data import get_site_configs, SITES_CONFIG_FOLDER, exists


#############################################
################# Functions #################
#############################################


def manage_interface_description(wb, osw_list, CMD_PATH):

    sheet = 'show_interface_description'

    ws = wb.create_sheet(title=sheet, index=0)

    print('Starting manage_interface_description')

    filename = [CMD_PATH + osw_list[0] + '_' + sheet + '.txt', CMD_PATH + osw_list[1] + '_' + sheet + '.txt']

    index_first_file = 0

    for file in filename:

        with open(file, 'r') as fin:
            if index_first_file == 0:
                myrow = 1
            else:
                myrow = index_first_file
            for line in fin:
                line = line.strip()
                line_list = line.split()
                for elem, mycol in zip(line_list, range(1, len(line_list) + 1)):
                    ws.cell(row=myrow, column=mycol, value=elem)
                myrow += 1
                index_first_file = myrow

    print('End manage_interface_description')


def manage_standby_brief(wb, osw_list, CMD_PATH):

    sheet = 'show_standby_brief'

    ws = wb.create_sheet(title=sheet, index=0)

    print('Starting manage_standby brief')

    filename = [CMD_PATH + osw_list[0] + '_' + sheet + '.txt', CMD_PATH + osw_list[1] + '_' + sheet + '.txt']
    index_first_file = 0

    for file in filename:

        with open(file, 'r') as fin:
            if index_first_file == 0:
                myrow = 1
            else:
                myrow = index_first_file
            for line in fin:
                line = line.strip()
                line_list = line.split()
                for elem, mycol in zip(line_list, range(1, len(line_list) + 1)):
                    ws.cell(row=myrow, column=mycol, value=elem)
                myrow += 1
                index_first_file = myrow

    print('End manage_standby brief')


def manage_vrrp_brief(wb, osw_list, CMD_PATH):

    sheet = 'show_vrrp_brief'

    ws = wb.create_sheet(title=sheet, index=0)

    print('Starting manage_vrrp_brief')

    filename = [CMD_PATH + osw_list[0] + '_' + sheet + '.txt', CMD_PATH + osw_list[1] + '_' + sheet + '.txt']
    index_first_file = 0

    for file in filename:

        with open(file, 'r') as fin:
            if index_first_file == 0:
                myrow = 1
            else:
                myrow = index_first_file
            for line in fin:
                line = line.strip()
                line_list = line.split()
                for elem, mycol in zip(line_list, range(1, len(line_list) + 1)):
                    ws.cell(row=myrow, column=mycol, value=elem)
                myrow += 1
                index_first_file = myrow

    print('End manage_vrrp_brief')


def manage_interface_trunk(wb, osw_list, CMD_PATH):

    possible_trunks = ['show_interface_po1_trunk',
                       'show_interface_po10_trunk',
                       'show_interface_po100_trunk']

    sheet = ''
    for sh in possible_trunks:
        filename = CMD_PATH + osw_list[0] + '_' + sh + '.txt'
        import os
        statinfo = os.stat(filename)
        if statinfo.st_size > 1:
            sheet = sh
            break

    ws = wb.create_sheet(title=sheet, index=0)
    text = ''
    lst = []
    vlan_count_dict = dict()

    print('Starting manage_interface_trunk')

    filename = [CMD_PATH + osw_list[0] + '_' + sheet + '.txt', CMD_PATH + osw_list[1] + '_' + sheet + '.txt']
    for file in filename:

        with open(file, 'r') as fin:
            text += fin.read()

    text_list = text.split('\n')
    first, second = get_indexes(text_list)

    for index in (first, second):

        _, vlan_string = text_list[index + 1].split()
        vlan_list = vlan_string.split(',')

        for v in vlan_list:
            if v.find('-') > 0:
                help_l = from_range_to_list(v)
                for elem in help_l:
                    lst.append(int(elem))
            else:
                lst.append(int(v))

    myset = set(lst)

    for set_elem in myset:
        vlan_count_dict[set_elem] = lst.count(set_elem)

    mycol = 1
    mykeys = list(vlan_count_dict.keys())
    mykeys.sort()
    max_row = len(mykeys) + 1

    for elem, myrow in zip(mykeys, range(1, max_row)):
        ws.cell(row=myrow, column=mycol, value=int(elem))
        ws.cell(row=myrow, column=mycol + 1, value=vlan_count_dict[elem])

    print('End manage_interface_trunk')


def manage_vlan_brief(wb, osw_list, CMD_PATH):

    sheet_osw1 = 'show_vlan_brief_OSW1'
    sheet_osw2 = 'show_vlan_brief_OSW2'
    sheet_osw1osw2 = 'show_vlan_brief'

    osw_vlanbrief_dict = dict()  # osw: fin.readlines()
    both_file_list = []

    ws_ows1 = wb.create_sheet(title=sheet_osw1, index=0)
    ws_osw2 = wb.create_sheet(title=sheet_osw2, index=0)
    myworksheet = [ws_ows1, ws_osw2]
    ws_osw1osw2 = wb.create_sheet(title=sheet_osw1osw2, index=0)

    print('Starting manage_vlan_brief')

    filename = [CMD_PATH + osw_list[0] + '_' + sheet_osw1osw2 + '.txt', CMD_PATH + osw_list[1] + '_' + sheet_osw1osw2 + '.txt']

    for file in filename:

        osw = file[len(CMD_PATH):-(5 + len(sheet_osw1osw2))]

        with open(file, 'r') as fin:
            osw_vlanbrief_dict[osw] = fin.readlines()

    for osw, ws in zip(osw_vlanbrief_dict.keys(), myworksheet):
        write_vlan_brief_on_sheet(ws, osw_vlanbrief_dict[osw])
        both_file_list += osw_vlanbrief_dict[osw]

    write_vlan_brief_on_sheet(ws_osw1osw2, both_file_list)

    print('End manage_vlan_brief')


def manage_nexus_vlan_db(wb, nexus_file_dict, OSW_LIST, AID_PATH):

    print('Starting manage_nexus_vlan_db')

    nexus_file_list = [ nexus_file_dict[OSW_LIST[0]][0],
                        nexus_file_dict[OSW_LIST[1]][0],
                        nexus_file_dict[OSW_LIST[0]][2],
                      ]

    nexusfile_to_sheet_dict = {nexus_file_list[0]: 'vlan_db_VCE1',
                               nexus_file_list[1]: 'vlan_db_VCE2',
                               nexus_file_list[2]: 'vlan_db_VSW'}
    vce_vlan_dict = dict()
    all_nexus_vlan = []

    for nexus_file in nexus_file_list:
        VCE_CFG_TXT_IN = AID_PATH + nexus_file
        vce_vlan_dict[nexus_file] = get_vlan_from_cfg(VCE_CFG_TXT_IN)
        ws = wb.create_sheet(title='VLAN_ON_{}'.format(nexusfile_to_sheet_dict[nexus_file]), index=0)
        myrow = 1
        for elem in vce_vlan_dict[nexus_file]:
            if elem:
                ws.cell(row=myrow, column=1, value=int(elem))

            myrow += 1
        all_nexus_vlan += vce_vlan_dict[nexus_file]

    all_nexus_vlan_int = [int(x) for x in all_nexus_vlan]
    ws = wb.create_sheet(title='VLAN_ON_ALL_NEXUS', index=0)
    myrow = 1
    all_nexus_vlan_set = set(all_nexus_vlan_int)
    all_nexus_vlan = list(all_nexus_vlan_set)
    all_nexus_vlan.sort()

    for elem in all_nexus_vlan:
        if elem:
            ws.cell(row=myrow, column=1, value=int(elem))
        myrow += 1

    print('End manage_nexus_vlan_db')


def manage_dot1q(wb, vce2vpe_po, nexus_file_dict, vpe_list, trunk_map, OSW_LIST, AID_PATH):

    # trunk_map = {vpe_node: [trunk1, trunk2, ]} where trunks are VPE to OSW trunks by VPE side
    print('Starting manage_dot1q on VCE and VPE')

    nexus_vceadd_file_list = [nexus_file_dict[OSW_LIST[0]][1],
                              nexus_file_dict[OSW_LIST[1]][1]]
    vceaddfile_to_sheet_dict = {nexus_vceadd_file_list[0]: 'dot1q_tag_on_VCE1',
                                nexus_vceadd_file_list[1]: 'dot1q_tag_on_VCE2'}

    vpe_file_list = [AID_PATH + vpe_list[0] + '.txt',
                     AID_PATH + vpe_list[1] + '.txt']
    vpefile_to_sheet_dict = {vpe_file_list[0]: 'dot1q_tag_on_VPE1',
                             vpe_file_list[1]: 'dot1q_tag_on_VPE2'}

    for nexus_file in vceaddfile_to_sheet_dict.keys():
        VCE_CFG_TXT_IN = AID_PATH + nexus_file
        ws = wb.create_sheet(title=vceaddfile_to_sheet_dict[nexus_file], index=0)
        create_sheet_for_vce_tag(ws, vce2vpe_po, VCE_CFG_TXT_IN)
    print('End manage_dot1q on VCE')

    for vpe_node, vpe_file in zip(trunk_map.keys(), vpe_file_list):
        ws = wb.create_sheet(title=vpefile_to_sheet_dict[vpe_file], index=0)
        create_sheet_for_vpe_tag(ws, vpe_node, vpe_file, trunk_map)


def manage_rb(wb, node_list, CMD_PATH):

    print('Starting  manage_rb')
    mac_osw_map = dict()  # {mac: name}
    vlan_rb_map = dict()  # {vlan: mac}
    new_vlan_rb_map = dict()
    ws = wb.create_sheet(title='Root-bridge_per_VLAN', index=0)

    for node in node_list:
        mac_osw_map[get_switch_mac_address(node, CMD_PATH)] = node
        vlan_rb_map.update(get_rb_per_vlan(node, CMD_PATH))

    for vlan in vlan_rb_map.keys():
        if vlan_rb_map[vlan] in mac_osw_map.keys():
            new_vlan_rb_map[int(vlan)] = mac_osw_map[vlan_rb_map[vlan]]
        else:
            new_vlan_rb_map[int(vlan)] = vlan_rb_map[vlan]

    myrow = 1
    for vlan in sorted(new_vlan_rb_map):
        ws.cell(row=myrow, column=1, value=int(vlan))
        ws.cell(row=myrow, column=2, value=new_vlan_rb_map[vlan])
        myrow += 1

    print('End  manage_rb')


def manage_static_routes(wb, vce_file_list):

    print('Starting  manage_static_routes')
    text_list = []
    ws = wb.create_sheet(title='VCE_Static_Routes', index=0)
    for file in vce_file_list:
        if 'VSW' not in file:
            parse1 = c.CiscoConfParse(file)
            routes_rough = parse1.find_lines(r'^ ip route')
            routes_rough_h = [x for x in routes_rough if 'ip router ospf' not in x]
            text_list += routes_rough_h
    myrow = 1

    for line in text_list:
        line_list = line.split()
        for elem, mycol in zip(line_list, range(1, len(line_list) + 1)):
            ws.cell(row=myrow, column=mycol, value=elem)
        myrow += 1
    print('End  manage_static_routes')

#######################################
########### Help Functions ############
#######################################


def get_rb_per_vlan(osw, CMD_PATH):
    ''' return a map {vlan: mac} indicating RB for osw '''

    sheet = 'show_spanning-tree_root_brief'

    file_name = CMD_PATH + osw + '_' + sheet + '.txt'

    show_list = from_file_to_cfg_as_list(file_name)

    mp = {}

    for elem in show_list:
        if len(elem) > 0:
            if elem[:2] == 'VL':
                lst_elem = elem.split()
                vlan = lst_elem[0]
                mac = lst_elem[2]
                mp[vlan[4:].lstrip('0')] = mac
            else:
                continue
        else:
            continue
    return mp


def get_switch_mac_address(osw, CMD_PATH):
    ''' return a string containing mac address of osw '''

    sheet = 'show_spanning-tree_bridge_address'

    file_name = CMD_PATH + osw + '_' + sheet + '.txt'

    lst = from_file_to_cfg_as_list(file_name)
    if lst is not None:
        mac = lst[1].split()[1]
    else:
        mac = None
    return mac


def from_file_to_cfg_as_list(file_name):
    ''' return a list containing text of file_name '''
    show_cmd = []

    for elem in open(file_name, 'r'):
        show_cmd.append(elem.rstrip())
    return show_cmd


def create_sheet_for_vpe_tag(ws, vpe_node, vpe_file, trunk_map):

    # trunk_map = {vpe_node: [trunk1, trunk2, ]} where trunks are VPE to OSW trunks by VPE side
    #for vpe_node, vpe_file in zip(trunk_map.keys(), vpe_file_list):
    parse_string = r''
    tag_list = []
    for trunk in trunk_map[vpe_node]:
        parse_string += '^interface {}\.|'.format(trunk)
    parse_string = parse_string[:-1]

    print("##########################")
    print(vpe_file)
    print(trunk_map)

    parse = c.CiscoConfParse(vpe_file)
    obj_list = parse.find_objects(parse_string)
    for obj in obj_list:
        if '.' in obj.text:
            list_line_obj = obj.text.split('.')
            tag = int(list_line_obj[-1])
            tag_list.append(tag)

    myrow = 1
    for tag in tag_list:
        ws.cell(row=myrow, column=1, value=int(tag))
        myrow += 1


def create_sheet_for_vce_tag(ws, intf, dev_file):
    ''' Returns a list of dot1q tag taken on VCE from "intf" trunk'''

    myrow = 1
    help_str = ''
    osw_vlan_set = set()
    parse = c.CiscoConfParse(dev_file)
    bepo_obj = parse.find_objects(r'^interface ' + intf)
    bepo_cfg = bepo_obj[0].ioscfg

    for line in bepo_cfg:
        s = line.split(' ')
        if len(s) >= 5:
            if s[3] == 'allowed' and s[5][0].isdigit():
                help_str += s[5] + ','
            elif s[3] == 'allowed' and s[5] == 'add':
                help_str += s[6] + ','

    help_list = help_str[:-1].split(',')

    for elem in help_list:
        if elem.find('-') > 0:
            help_l = from_range_to_list(elem)
            for elemh in help_l:
                osw_vlan_set.add(int(elemh))
        else:
            osw_vlan_set.add(int(elem))

    result_l = list(osw_vlan_set)
    result_l.sort()

    for tag in result_l:
        ws.cell(row=myrow, column=1, value=int(tag))
        myrow += 1


def get_vlan_from_cfg(filepath):
    ''' get vlan from VCE config '''

    vlan_list = []
    parse1 = c.CiscoConfParse(filepath)
    vlan_rough = parse1.find_lines(r'^vlan')

    for v in vlan_rough:
        vlan_list.append(v.split()[1])

    return vlan_list


def write_vlan_brief_on_sheet(ws, vlan_brief_list):

    myrow = 1

    for line in vlan_brief_list:
        if len(line) > 1:
            if line[:4] != 'VLAN' and line[:4] != '----' and line[:4] != 'show':
                line_list = line.split()
                if line_list[0][0].isnumeric() or line_list[0] == 'show':
                    for elem, mycol in zip(line_list, range(1, len(line_list) + 1)):
                        if mycol == 1:
                            ws.cell(row=myrow, column=mycol, value=int(elem))
                        else:
                            ws.cell(row=myrow, column=mycol, value=elem)
                    myrow += 1
                else:
                    continue
            else:
                continue
        else:
            continue


def from_range_to_list(range_str):
    ''' tansform '1-3' in [1,2,3] '''

    mylist = []

    h_l = range_str.split('-')
    start = int(h_l[0])
    stop = int(h_l[1])
    for x in range(start, stop + 1):
        mylist.append(x)
    return mylist


def get_indexes(text_list):

    bool_line_match = False
    for line in text_list:
        line = line.strip()
        if line == 'Port          Vlans allowed on trunk' and bool_line_match is False:
            first_index = text_list.index(line)
            bool_line_match = True
        elif line == 'Port          Vlans allowed on trunk' and bool_line_match is True:
            second_index = text_list.index(line, text_list.index(line) + 1)

    return (first_index, second_index)



def copy_folder(site_configs):

    for site_config in site_configs:
        #copying site config
        source_path = site_config.base_dir + site_config.site + "FINAL/"
        source_file_osw = source_path + site_config.switch + "VCE.txt"
        source_file_vsw = source_path + site_config.switch + "VSW.txt"
        source_path = site_config.base_dir + site_config.site + "DATA_SRC/CFG/"
        source_file_vpe = source_path + site_config.vpe_router + ".txt"
        dest_path = site_config.base_dir + site_config.site + "AID/"
        dest_file_osw = dest_path + site_config.switch + "VCE.txt"
        dest_file_vsw = dest_path + site_config.switch + "VSW.txt"
        dest_file_vpe = dest_path + site_config.vpe_router + ".txt"
        for dest_file, source_file in zip([dest_file_osw, dest_file_vsw, dest_file_vpe],
                                          [source_file_osw, source_file_vsw, source_file_vpe]):
            if exists(dest_file):
                print(dest_file + " already exists.")
            else:
                print("Copying " + dest_file)
                copy_file(source_file, dest_file, dest_path)

    for site_config in site_configs:
        #copying site config
        source_path = site_config.base_dir + site_config.site + "FINAL/"
        source_file_osw = source_path + site_config.switch + "VCE_addendum.txt"
        source_file_vsw = source_path + site_config.switch + "VPE_addendum.txt"
        dest_path = site_config.base_dir + site_config.site + "AID/"
        dest_file_osw = dest_path + site_config.switch + "VCE_addendum.txt"
        dest_file_vsw = dest_path + site_config.switch + "VPE_addendum.txt"
        for dest_file, source_file in zip([dest_file_osw, dest_file_vsw], [source_file_osw, source_file_vsw]):
            if exists(dest_file):
                print(dest_file + " already exists.")
            else:
                print("Copying " + dest_file)
                copy_file(source_file, dest_file, dest_path)

def create_dir(dest_path):
    import os
    if not os.path.exists(dest_path):
        os.makedirs(dest_path)


def copy_file(source_file, dest_file, dest_path):
    import shutil
    if not exists(source_file):
        print("File " + source_file + " is missing. \nPlease create it.")
        exit(0)
    create_dir(dest_path)
    shutil.copy(source_file, dest_file)
############################################
################# MAIN #####################
############################################


def run(site_configs):
    import re

    OSW2OSW_PO = 'Port-channel' + site_configs[0].portch_OSW_OSW
    VCE2VPE_PO = 'Port-channel' + site_configs[0].portch_VCE_VPE

    CMD_PATH = site_configs[0].base_dir + site_configs[0].site + "/DATA_SRC/CMD/"
    AID_PATH = site_configs[0].base_dir + site_configs[0].site  + 'AID/'

    site = site_configs[0].switch
    m = re.search('([A-Z]{2})OSW(\d\d)', site)
    site = m.group(1) + m.group(2)

    OUTPUT_XLS = AID_PATH + 'AID_to_{}_NMP.xlsx'.format(site)

    OSW_LIST = []
    VPE_LIST = []
    VPE_FILE_LIST = []
    VCE_FILE_LIST = []
    NEXUS_FILE_DICT = {}
    TRUNK_MAP = {}
    for i in range(0, 2):
        OSW_LIST.append(site_configs[i].switch)
        VPE_LIST.append(site_configs[i].vpe_router)

        # trunk_map = {vpe_node: [trunk1, trunk2, ]} where trunks are VPE to OSW trunks by VPE side
        if len(list(site_configs[i].vpeosw_to_vpevce.keys())) > 0:
            TRUNK_MAP[VPE_LIST[i]] = ['Bundle-Ether' + site_configs[i].portch_OSW_VPE,
                                      list(site_configs[i].vpeosw_to_vpevce.keys())]

        NEXUS_FILE_DICT[OSW_LIST[i]] = [OSW_LIST[i] + 'VCE.txt',
                                         OSW_LIST[i] + 'VCE_addendum.txt',
                                         OSW_LIST[i] + 'VSW.txt',
                                         VPE_LIST[i] + 'VPE_addendum.txt'
                                       ]

        VPE_FILE_LIST.append(AID_PATH + VPE_LIST[i] + '.txt')
    
        VCE_FILE_LIST.append(AID_PATH + NEXUS_FILE_DICT[OSW_LIST[i]][0])


    wb = Workbook()
    manage_interface_description(wb, OSW_LIST, CMD_PATH)
    manage_standby_brief(wb, OSW_LIST, CMD_PATH)
    manage_vrrp_brief(wb, OSW_LIST, CMD_PATH)
    manage_interface_trunk(wb, OSW_LIST, CMD_PATH)
    manage_vlan_brief(wb, OSW_LIST, CMD_PATH)
    manage_nexus_vlan_db(wb, NEXUS_FILE_DICT, OSW_LIST, AID_PATH)
    manage_dot1q(wb, VCE2VPE_PO, NEXUS_FILE_DICT, VPE_LIST, TRUNK_MAP, OSW_LIST, AID_PATH)
    manage_rb(wb, OSW_LIST, CMD_PATH)
    manage_static_routes(wb, VCE_FILE_LIST)
    wb.save(filename=OUTPUT_XLS)

if __name__ == "__main__":
    site_configs = get_site_configs(SITES_CONFIG_FOLDER)
    copy_folder(site_configs)
    run(site_configs)