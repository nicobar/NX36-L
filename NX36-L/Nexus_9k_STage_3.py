from openpyxl import load_workbook
import ciscoconfparse as c
import re
import ipaddress
import itertools


######################################################
################# VARIABLES/CONSTANT #################
######################################################

SWITCH = 'BOOSW013'

SHEET = SWITCH

VLAN_FROM_XLS = True  # IF TRUE THEN XLS IS TRUSTABLE AS TRUSTED VLAN' SOURCE

BASE = '/mnt/hgfs/VM_shared/VF-2017/NMP/'
SITE = 'BO01/'
BASE_DIR = BASE + SITE + SWITCH + '/Stage_3/'

INPUT_XLS = BASE_DIR + SWITCH + '_OUT_DB_OPT.xlsx'
OSW_CFG_TXT = BASE_DIR + SWITCH + '.txt'
OSWVCE_CFG_TXT = BASE_DIR + SWITCH + 'VCE' + '.txt'
OSWVSW_CFG_TXT = BASE_DIR + SWITCH + 'VSW' + '.txt'

OTHER_SWITCH = 'BOOSW016'


OTHER_BASE_DIR = BASE + SITE + OTHER_SWITCH + '/Stage_3/'
OTHER_INPUT_XLS = OTHER_BASE_DIR + OTHER_SWITCH + '_OUT_DB_OPT.xlsx'
OTHER_SHEET = OTHER_SWITCH

PO_OSW_OSW = r'^interface Port-channel1$'
PO_OSW_VPE = r'^interface Port-channel113$'

qos_sp_def_N9508_dict = {
    'U': ' service-policy type qos input UNTRUST',
    'T': ' service-policy type qos input VF-INGRESS',
    'S': ' service-policy type qos input SIGNALLING',
    'V': ' service-policy type qos input VOICE',
    'D': ' service-policy type qos input 2G_3G_DATA',
    'K': ' service-policy type qos input POLICY_MGW'}

qos_sp_def_N3048_dict = {
    'U': ' service-policy type qos input UNTRUST',
    'T': '',
    'S': ' service-policy type qos input SIGNALLING',
    'V': ' service-policy type qos input VOICE',
    'D': ' service-policy type qos input 2G_3G_DATA'}

#############################################
################ FUNCTIONS ##################
#############################################


# set_a | set_b --> Union
# set_a & set_b --> intersection

def atoi(text):
    ''' from string to int'''

    return int(text) if text.isdigit() else text


def natural_keys(text):
    '''
    alist.sort(key=natural_keys) sorts in human order
    http://nedbatchelder.com/blog/200712/human_sorting.html
    (See Toothy's implementation in the comments)
    '''

    return [atoi(c) for c in re.split('(\d+)', text)]


def get_ordered_list(main_lst):
    '''
    Get a list of string ["95", "31,32","10-13", "1,3-5"] and return a
    list of string  ['1', '3', '4', '5', '10', '11', '12',
    '13', '31', '32', '95']
    '''

    a_set = set()

    for elem1 in main_lst:
        if ',' in elem1 and '-' not in elem1:
            elem1 = elem1.split(',')
            for elem2 in elem1:
                a_set.add(elem2)
        elif ',' not in elem1 and '-' in elem1:
            lst = from_range_to_list(elem1)
            for elem2 in lst:
                a_set.add(elem2)
        elif ',' in elem1 and '-' in elem1:
            elem1 = elem1.split(',')
            for elem2 in elem1:
                if '-' in elem2:
                    lst = from_range_to_list(elem2)
                    for elem3 in lst:
                        a_set.add(elem3)
                else:
                    a_set.add(elem2)
        else:
            a_set.add(elem1)
    a = list(a_set)
    a.sort(key=natural_keys)
    return a


def get_col_N3048(ws, col):
    ''' Take a worksheet, return column "col" as lists conditioned to col = 6 == "N3048" '''

    NEXUS_AP_COL = 6
    return [str(ws.cell(row=r, column=col).value) for r in range(2, ws.max_row)
            if ws.cell(row=r, column=NEXUS_AP_COL).value == 'N3048']


def get_col_for_vlan_N9508(ws, col):
    ''' Take a worksheet, return column "col" as lists conditioned to col = 6 != "N3048" '''

    NEXUS_AP_COL = 6
    return [str(ws.cell(row=r, column=col).value) for r in range(2, ws.max_row + 1)
            if ws.cell(row=r, column=NEXUS_AP_COL).value != 'N3048']


def get_col_N9508(ws, col):
    ''' Take a worksheet, return column "col" as lists conditioned to col = 6 != "N3048" '''

    NEXUS_AP_COL = 6
    return [str(ws.cell(row=r, column=col).value) for r in range(2, ws.max_row + 1)
            if (ws.cell(row=r, column=NEXUS_AP_COL).value != 'N3048' and ws.cell(row=r, column=NEXUS_AP_COL).value != 'Infra')]


def get_if_from_xls():
    ''' Return column col as list '''

    wb_r = load_workbook(INPUT_XLS)
    ws_r = wb_r.get_sheet_by_name(SHEET)
    SRC_IF_COL = 1
    if_N3048 = get_col_N3048(ws_r, SRC_IF_COL)
    if_N9508 = get_col_N9508(ws_r, SRC_IF_COL)
    if_N3048.sort(key=natural_keys)
    if_N9508.sort(key=natural_keys)
    return (if_N9508, if_N3048)


def get_if_from_cfg():
    ''' from cfg get list of all *Ethernet interface '''

    parse = c.CiscoConfParse(OSW_CFG_TXT)
    intf_obj_list = parse.find_objects(r'^interface .*Ethernet')

    a = [obj.text for obj in intf_obj_list]
    a.sort(key=natural_keys)
    return a


def get_vlan_from_cfg():
    ''' from cfg get list of all L2 vlan  '''

    parse = c.CiscoConfParse(OSW_CFG_TXT)
    vlan_obj_list = parse.find_objects(r'^vlan \d+')

    return [obj.text.split(' ')[1] for obj in vlan_obj_list]


def get_vlan_from_xls():
    ''' from xls get list of all L2 vlan  '''

    wb_r = load_workbook(INPUT_XLS)
    ws_r = wb_r.get_sheet_by_name(SHEET)
    VLAN_COL = 4

    lst_N9508 = get_col_for_vlan_N9508(ws_r, VLAN_COL)
    lst_N3048 = get_col_N3048(ws_r, VLAN_COL)

    lst_N9K = get_ordered_list(lst_N9508)
    lst_N3K = get_ordered_list(lst_N3048)

    return (lst_N9K, lst_N3K)


def get_vlan_other_sw_from_xls():
    ''' from xls get list of all L2 vlan  '''

    a = set()

    wb_r = load_workbook(OTHER_INPUT_XLS)
    ws_r = wb_r.get_sheet_by_name(OTHER_SHEET)
    VLAN_COL = 4

    lst_N9508 = get_col_for_vlan_N9508(ws_r, VLAN_COL)
    lst_N3048 = get_col_N3048(ws_r, VLAN_COL)
    lst1 = lst_N9508 + lst_N3048
    st = set(lst1)
    lst = list(st)
    lst.sort(key=natural_keys)

    for elem in lst:
        if ',' in elem:
            b = elem.split(',')
            for elem2 in b:
                a.add(elem2)
        else:
            a.add(elem)

    lst2 = list(a)

    lst2.sort(key=natural_keys)

    return lst2


def get_svi_from_cfg():
    ''' from cfg get list of all svi interfaces  '''

    parse = c.CiscoConfParse(OSW_CFG_TXT)
    svi_obj_list = parse.find_objects(r'^interface Vlan')

    lst = [re.findall(r'^interface Vlan(\d+)', svi_obj.text)[0]
           for svi_obj in svi_obj_list]
    return lst


def get_svi_on_device(vlanxls, svi_from_cfg):
    ''' creates a list of svi interface that are both in cfg as svi and in xls as vlan '''

    a = [x for x in svi_from_cfg if x in vlanxls]
    a.sort(key=natural_keys)
    return a


def get_list_not_to_be_migrated(ifxls, ifcfg):
    ''' given set(a) and set(b), returns list(b-a) if b > a else []  '''

    a = set(ifxls)
    b = set(ifcfg)
    c = b - a
    d = list(c)
    if len(d) > 0:
        d.sort(key=natural_keys)
        return d
    else:
        return []


def get_migration_dictionary_N3048():
    ''' return {SRC_IF:DEST_IF} for N3048  '''

    wb_r = load_workbook(INPUT_XLS)
    ws_r = wb_r.get_sheet_by_name(SHEET)

    NEXUS_AP_COL = 6
    SRC_IF_COL = 1
    DST_IF_COL = 2
    return {str(ws_r.cell(row=r, column=SRC_IF_COL).value): str(ws_r.cell(row=r, column=DST_IF_COL).value)
            for r in range(2, ws_r.max_row + 1) if ws_r.cell(row=r, column=NEXUS_AP_COL).value == 'N3048'}


def get_migration_dictionary_N9508():
    ''' return {SRC_IF:DEST_IF} for N9508  '''

    wb_r = load_workbook(INPUT_XLS)
    ws_r = wb_r.get_sheet_by_name(SHEET)

    NEXUS_AP_COL = 6
    SRC_IF_COL = 1
    DST_IF_COL = 2
    return {str(ws_r.cell(row=r, column=SRC_IF_COL).value): str(ws_r.cell(row=r, column=DST_IF_COL).value)
            for r in range(2, ws_r.max_row + 1) if (ws_r.cell(row=r, column=NEXUS_AP_COL).value != 'N3048' and ws_r.cell(row=r, column=NEXUS_AP_COL).value != 'Infra')}


def get_normalized_if_OSWVCEVSW_cfg(if_ntbm, mig_dict, qos_sp_def_dict):
    ''' return cfg as list of migrated and cleaned - fz clean_if_cfg - interfaces  '''

    intf_gr = get_if_xls_guardroot()

    intf_qos_dict = get_if_to_qos_xls_dict()
    parse = c.CiscoConfParse(OSW_CFG_TXT)

    intf_obj_list = parse.find_objects(r'^interface .*Ethernet')

    for intf_obj in intf_obj_list:
        if intf_obj.text in if_ntbm:
            intf_obj.delete()
        elif intf_obj.text in mig_dict:
            intf_obj.replace(intf_obj.text, mig_dict[intf_obj.text])
            if intf_obj.text in intf_gr:
                intf_obj.append_to_family(
                    'spanning-tree guard root', auto_indent=True)
            if intf_obj.text in intf_qos_dict:
                intf_obj.append_to_family(
                    qos_sp_def_dict[intf_qos_dict[intf_obj.text]], auto_indent=True)

    parse.commit()

    intf_obj_list = parse.find_objects(r'^interface .*Ethernet')
    cf_intf_list = [intf_obj.ioscfg + ['!'] for intf_obj in intf_obj_list]
    cf_intf_1 = list(itertools.chain.from_iterable(cf_intf_list))
    cf_intf_2 = clean_if_cfg(cf_intf_1)
    cf_intf_3 = add_nonegotiationauto(cf_intf_2)
    cf_intf = add_shutdown(cf_intf_3)
    return cf_intf


def get_normalized_vlan_OSWVCEVSW_cfg(vlan_ntbm):
    ''' return cfg as list of vlans  '''

    parse = c.CiscoConfParse(OSW_CFG_TXT)

    vlan_obj_list = parse.find_objects(r'^vlan \d+$')

    for vlan_obj in vlan_obj_list:
        vlan = vlan_obj.text
        if re.findall(r'^vlan (\d+)$', vlan)[0] in vlan_ntbm:
            vlan_obj.delete()

    parse.commit()
    vlan_obj_list = parse.find_objects(r'^vlan \d+$')
    cf_vlan_list = [vlan_obj.ioscfg + ['!'] for vlan_obj in vlan_obj_list]
    cf_vlan = list(itertools.chain.from_iterable(cf_vlan_list))
    return cf_vlan

#def get_normalized_svi_OSWVCEVSW_cfg(svi_ntbm, svi_on_device):


def get_normalized_svi_OSWVCEVSW_cfg(svi_ntbm):
    ''' return cfg as list of svi  '''

    parse = c.CiscoConfParse(OSW_CFG_TXT)
    svi_obj_list = parse.find_objects(r'^interface Vlan')

    for svi_obj in svi_obj_list:
        svi = svi_obj.text
        num_svi = re.findall(r'^interface Vlan(\d+)$', svi)[0]
        if num_svi in svi_ntbm:
            svi_obj.delete()

    parse.commit()
    svi_obj_list = parse.find_objects(r'^interface Vlan')
    cf_svi_list = [svi_obj.ioscfg + ['!'] for svi_obj in svi_obj_list]
    cf_svi_1 = list(itertools.chain.from_iterable(cf_svi_list))
    #cf_svi_2 = clean_hsrp_to_svi(svi_on_device, cf_svi_1)
    cf_svi_2 = clean_hsrp_to_svi(cf_svi_1)
    cf_svi = add_shutdown(cf_svi_2)
    return cf_svi


def clean_if_cfg(cfg):
    ''' cleans if cfg (lst) with command stored in  command_to_be_deleted set '''

    parse = c.CiscoConfParse(cfg)
    command_to_be_deleted = (
        r'no cdp enable',
        r'spanning-tree bpduguard enable',
        r'no ip address',
        r'switchport trunk encapsulation dot1q',
        r'speed nonegotiate')

    intf_obj_list = parse.find_objects(r'^interface')
    for intf_obj in intf_obj_list:
        for c_tbd in command_to_be_deleted:
            intf_obj.delete_children_matching(c_tbd)
    parse.commit()
    intf_obj_list = parse.find_objects(r'^interface')
    cf_intf_list = [intf_obj.ioscfg + ['!'] for intf_obj in intf_obj_list]
    cf_intf = list(itertools.chain.from_iterable(cf_intf_list))
    return cf_intf


def add_nonegotiationauto(cfg):
    ''' add to if cfg (lst) with "no auto negotiation" command if speed is present '''

    parse = c.CiscoConfParse(cfg)

    parse.insert_after_child(
        parentspec=r"^interface Ether",
        childspec=r"speed 100",
        insertstr=' no negotiate auto')
    parse.commit()
    return parse.ioscfg

#def clean_hsrp_to_svi(svi_tbm_list, cfg):


def clean_hsrp_to_svi(cfg):
    ''' translate HSRP from IOS to NS-OX '''

    config_svi_to_add = []
    parse = c.CiscoConfParse(cfg)

    svi_hsrp_obj_list = parse.find_objects(r'^interface Vlan')

    svi_hsrp_list_h2 = [obj.ioscfg for obj in svi_hsrp_obj_list]
    svi_hsrp_list_h1 = list(itertools.chain.from_iterable(svi_hsrp_list_h2))

    # eliminates VRRP commands (will become HSRP secondary on VPE)
    svi_hsrp_list = [svi for svi in svi_hsrp_list_h1 if svi.find('vrrp') < 0]

    for line in svi_hsrp_list:
        line_lst = line.lstrip().split()

        if line_lst[0] == 'standby':
            if len(line_lst) == 4:
                if line_lst[2] == 'ip':
                    config_svi_to_add.append(' hsrp ' + line_lst[1])
                    config_svi_to_add.append('  ip ' + line_lst[3])
                elif line_lst[2] == 'priority':
                    config_svi_to_add.append('  priority ' + line_lst[3])
            elif len(line_lst) == 3:
                if line_lst[2] == 'preempt':
                    config_svi_to_add.append('  preempt')
                else:
                    config_svi_to_add.append('  ERROR TO BE CHECKED ')
        elif line_lst[0] == 'interface':
            config_svi_to_add.append('!')
            config_svi_to_add.append(line)
        else:
            config_svi_to_add.append(line)

    return config_svi_to_add


def add_shutdown(cfg):
    ''' add  shutdown after 'interface' to list 'cfg'  '''

    new_cfg = []
    for line in cfg:
        if line.lstrip().split()[0] == 'interface':
            new_cfg.append(line)
            new_cfg.append(' shutdown')
        else:
            new_cfg.append(line)
    return new_cfg


def get_cleaned_routes():  # return a list of cleaned (with egress interfaces) routes
    ''' return a list with all OSW routes with the right egress interface  '''

    new_route_list = ['!', 'vrf context OPNET']
    parse = c.CiscoConfParse(OSW_CFG_TXT)

    route_list = parse.find_lines(r'^ip route')

    parse.commit()

    l3_obj_list = parse.find_objects_w_child(r'^interface', r'^ ip address')
    for route in route_list:
        route_l = route.split()
        nh = route_l[4]
        if nh[0].isdigit(
        ):  # interested in those routes that do not have exit interface in statement
            ip_nh = ipaddress.IPv4Address(nh)
            for l3_obj in l3_obj_list:
                if l3_obj.text != 'interface Loopback0':
                    for elem in l3_obj.ioscfg:
                        if elem[:11] == ' ip address':
                            help_list = elem.split()
                            ip_network = ipaddress.IPv4Network(
                                help_list[2] + '/' + help_list[3], strict=False)
                            if ip_nh in ip_network:
                                nh_ifs = l3_obj.text[10:]
                                route_l.insert(route_l.index(nh), nh_ifs)
                                new_route = '  ' + ' '.join(route_l)

        else:

            new_route = '  ' + route
        new_route_list.append(new_route)

    return new_route_list + ['!']


# def get_svi_with_static_on_N3048(static_routes, svi_on_vsw):
#     ''' return a list of svi that are egress intf in static routes  '''
#
#     real_svi_on_vsw = []
#     for route in static_routes:
#         svi = route.split(' ')
#         if len(svi) >= 6:
#             out_svi = re.findall(r'(\d+)', svi[6])[0]
#             if out_svi in svi_on_vsw:
#                 real_svi_on_vsw.append(out_svi)
#
#     return real_svi_on_vsw


def get_routes_for_devices(static_routes, real_svi_on_device):
    ''' return a list of routes whose egress if are referenced in static routes   '''

    real_routes_on_device = ['\n' + 10 * '#' + " 6500's routes here below " + 1 * '#' + '\n']
    for route in static_routes:
        svi = route.split(' ')
        if len(svi) >= 6:
            out_svi = re.findall(r'(\d+)', svi[6])[0]
            if out_svi in real_svi_on_device:
                real_routes_on_device.append(route)

    return real_routes_on_device + ['!']


def get_net_in_area():
    ''' return a dict of kind: { area: [ipaddress('prefix/lenght'),] } '''

    dict_net_in_area = dict()  # { area: [ipaddress('prefix/lenght'),] }

    parse = c.CiscoConfParse(OSW_CFG_TXT)

    ospf_obj_list = parse.find_objects(r'^router ospf')

    for line in ospf_obj_list[0].ioscfg:
        hl = line.split()
        if hl[0] == 'network':
            net = hl[1]
            net_hostmask = hl[2]
            area = hl[4]
            if area not in dict_net_in_area:
                dict_net_in_area[area] = [
                    ipaddress.IPv4Network(
                        net + '/' + net_hostmask)]
            else:
                dict_net_in_area[area].append(
                    ipaddress.IPv4Network(
                        net + '/' + net_hostmask))
    for area in dict_net_in_area:
        dict_net_in_area[area] = list(
            ipaddress.collapse_addresses(
                dict_net_in_area[area]))
    return dict_net_in_area


def get_svi_to_area(d_net_in_area):
    ''' take { area: [ipaddress('prefix/lenght'),] } and returns a dict of kind: { SVI : area } '''

    dict_svi_to_area = dict()  # { SVI : area }
    parse = c.CiscoConfParse(OSW_CFG_TXT)

    svi_obj_list = parse.find_objects(r'^interface Vlan')

    for svi_obj in svi_obj_list:
        for line in svi_obj.ioscfg:
            hl = line.split()
            if hl[0] == 'ip' and hl[1] == 'address':
                ip_svi = ipaddress.IPv4Address(hl[2])
                dict_svi_to_area[svi_obj.text] = -1
                for area in d_net_in_area:
                    for net in d_net_in_area[area]:
                        if ip_svi in net:
                            if int(area) <= 255:
                                dict_svi_to_area[svi_obj.text] = '0.0.0.' + area
                            else:
                                dict_svi_to_area[svi_obj.text] = area

    return dict_svi_to_area


def add_ospf_to_svi_cfg(svi_conf_list, svi_on_device, d_svi_to_area):
    ''' take svi cfg and returns a the same cfg with ospf part '''

    svi_with_ospf_conf = []
    vlan_svi = None

    for line in svi_conf_list:

        if re.match('interface Vlan', line):
            vlan_svi = re.findall(r'\d+', line)[0]
            int_svi = line

        if vlan_svi in svi_on_device:
            line_lst = line.lstrip().split()
            if re.match('interface Vlan', line):
                # svi_with_ospf_conf.append('!')
                svi_with_ospf_conf.append(line)
            elif line_lst[0] == 'ip' and line_lst[1] == 'address':
                if int_svi in d_svi_to_area:
                    if isinstance(
                            d_svi_to_area[int_svi], str) and d_svi_to_area[int_svi] != '':

                        svi_with_ospf_conf.append(' vrf member OPNET')
                        svi_with_ospf_conf.append(line)
                        svi_with_ospf_conf.append(
                            ' ip router ospf 249 area ' + str(d_svi_to_area['interface Vlan' + vlan_svi]))

                    elif d_svi_to_area[int_svi] == '' or isinstance(d_svi_to_area[int_svi], str) is False:
                        svi_with_ospf_conf.append(' vrf member OPNET')
                        svi_with_ospf_conf.append(line)

                else:
                    svi_with_ospf_conf.append(' vrf member OPNET')
                    svi_with_ospf_conf.append(line)

            else:
                svi_with_ospf_conf.append(line)

    return svi_with_ospf_conf


def get_if_xls_guardroot():
    ''' Return intf list if root_guard == 'Yes' '''

    wb_r = load_workbook(INPUT_XLS)
    ws_r = wb_r.get_sheet_by_name(SHEET)
    DST_VCE_IF_COL = 2
    ROOT_GUARD_COL = 13

    if_gr = [str(ws_r.cell(row=r, column=DST_VCE_IF_COL).value) for r in range(
        2, ws_r.max_row + 1) if str(ws_r.cell(row=r, column=ROOT_GUARD_COL).value) == 'Yes']
    if_gr.sort(key=natural_keys)
    return if_gr


def get_if_to_qos_xls_dict():
    ''' Return {intf : QoS} dict '''

    wb_r = load_workbook(INPUT_XLS)
    ws_r = wb_r.get_sheet_by_name(SHEET)
    DST_VCE_IF_COL = 2
    QOS_COL = 5
    NEXUS_AP = 6

    qos_gr = {str(ws_r.cell(row=r, column=DST_VCE_IF_COL).value): str(ws_r.cell(row=r, column=QOS_COL).value)
              for r in range(2, ws_r.max_row + 1) if str(ws_r.cell(row=r, column=NEXUS_AP).value) != 'Infra'}

    return qos_gr


def from_range_to_list(range_str):

    mylist = []

    h_l = range_str.split('-')
    start = int(h_l[0])
    stop = int(h_l[1])
    for x in range(start, stop + 1):
        mylist.append(str(x))
    return mylist


def get_vlan_list_from_po(po):

    po_vlan1 = []

    parse = c.CiscoConfParse(OSW_CFG_TXT)
    intf_obj = parse.find_objects(po)
    spur_list = [
        re.findall(
            r'switchport trunk allowed vlan *[add]* ([\d+,-]*)',
            x) for x in intf_obj[0].ioscfg]
    spur_list0 = list(itertools.chain.from_iterable(spur_list))
    spur_list1 = [x.split(',') for x in spur_list0]
    spur_list2 = list(itertools.chain.from_iterable(spur_list1))

    for y in spur_list2:
        if '-' in y:
            m = from_range_to_list(y)
            for x in m:
                po_vlan1.append(x)
        else:
            po_vlan1.append(y)

    po_vlan_set = set(po_vlan1)
    po_vlan = list(po_vlan_set)
    po_vlan.sort(key=natural_keys)
    return po_vlan


def get_vlan_to_add():

    #list_a = get_vlan_from_xls()
    list_b = get_vlan_list_from_po(PO_OSW_VPE)
    list_c = get_vlan_list_from_po(PO_OSW_OSW)
    list_other = get_vlan_other_sw_from_xls()

    print("Allowed VLAN list on OSW<-->VPE Port-channel : ", list_b)
    print("Allowed VLAN list on OSW<-->OSW Port-channel : ", list_c)
    #set_a = set(list_a)
    set_b = set(list_b)
    set_c = set(list_c)

    # set_a | set_b --> Union
    # set_a & set_b --> intersection
    #
    # v_to_add = list(set_a | (set_b & set_c))
    v_to_add = list(set_b & set_c)
    v_to_add.sort(key=natural_keys)
    # to have just VLAN in both infrastructure trunks AND on other switch in
    # access/trunk if
    vv_to_add = [x for x in v_to_add if x in list_other]

    print("VLAN list MUST BE PRESENT on this OSW withouth access/trunk if", vv_to_add)

    return vv_to_add


def transform_routes_for_nexus(routes):
    '''        0   1     2     3       4     5  6   7      8      9      10
        from <ip route prefix mask interface nh ad tag tag_value name name_value>
        to   <ip route prefix mask interface nh name name_value tag tag_value ad>
    '''
    new_route_list = []
    n9k_routes = ['!', 'vrf context OPNET']
    for route in routes[1:]:  # first line is a speratore string made of #
        route_list = route.split()
        if len(route_list) == 11:  # complete
            new_route_list = route_list[:6] + [' '.join(route_list[-2:])] + [' '.join(route_list[-4:-2])] + [route_list[6]]
        elif len(route_list) == 10:  # ad missed
            new_route_list = route_list[:6] + [' '.join(route_list[-2:])] + [' '.join(route_list[-4:-2])]
        elif len(route_list) == 9:  # tag or name missing
            if route_list[7] == 'tag':
                new_route_list = route_list[:6] + [' '.join(route_list[-2:])] + [route_list[6]]
            elif route_list[7] == 'name':
                new_route_list = route_list[:6] + [' '.join(route_list[-2:])] + [route_list[6]]
        elif len(route_list) == 8:  # ad and tag or name missing
            new_route_list = route_list
        elif len(route_list) == 7:  # tag and name missing
            new_route_list = route_list
        elif len(route_list) == 6:  # ad tag and name missing
            new_route_list = route_list
        elif len(route_list) == 3 and route_list[0] == 'vrf':
            continue
        elif len(route_list) == 1 and route_list[0] == '!':
            continue
        else:
            new_route_list = ['ERRORE']

        new_route_string = ' '.join(new_route_list)
        n9k_routes.append(' ' + new_route_string)

    return n9k_routes


#############################################
################### MAIN ####################
#############################################

############## ROUTES ###############


routes = get_cleaned_routes()
for r in routes:
    print(r)


########### IF #########

if_xls_N9508, if_xls_N3048 = get_if_from_xls()
if_cfg = get_if_from_cfg()

print("if_xls_N9508 = ", if_xls_N9508)
print("if_xls_N3048 = ", if_xls_N3048)
print("if_cfg = ", if_cfg)

if_not_to_be_migrated_N9508 = get_list_not_to_be_migrated(if_xls_N9508, if_cfg)
if_not_to_be_migrated_N3048 = get_list_not_to_be_migrated(if_xls_N3048, if_cfg)
print("if_not_to_be_migrated_N9508 = ", if_not_to_be_migrated_N9508)
print("if_not_to_be_migrated_N3048 = ", if_not_to_be_migrated_N3048)

################ VLAN CANDIDATE ##############


candidate_vlan_xls_N9508, candidate_vlan_xls_N3048 = get_vlan_from_xls()
candidate_vlan_xls_N9508 = list(set(candidate_vlan_xls_N9508) | set(candidate_vlan_xls_N3048))
candidate_vlan_xls_N9508.sort(key=natural_keys)
vlan_cfg = get_vlan_from_cfg()  # all vlan from CFG, even those NTBM

print("candidate_vlan_xls_N9508 = ", candidate_vlan_xls_N9508)
print("candidate_vlan_xls_N3048 = ", candidate_vlan_xls_N3048)


# IF TRUE THEN XLS IS TRUSTABLE AS TRUSTED VLAN SOURCE; Here we test False case
if not VLAN_FROM_XLS:
    vlan_to_add = get_vlan_to_add()
    candidate_vlan_xls_N9508 += vlan_to_add
    candidate_vlan_xls_N9508.sort(key=natural_keys)


candidate_vlan_not_to_be_migrated_N9508 = get_list_not_to_be_migrated(candidate_vlan_xls_N9508, vlan_cfg)
candidate_vlan_not_to_be_migrated_N3048 = get_list_not_to_be_migrated(candidate_vlan_xls_N3048, vlan_cfg)
print("candidate_vlan_not_to_be_migrated_N9508 = ", candidate_vlan_not_to_be_migrated_N9508)
print("candidate_vlan_not_to_be_migrated_N3048 = ", candidate_vlan_not_to_be_migrated_N3048)

############## SVI ################

svi_from_cfg = get_svi_from_cfg()
candidate_svi_on_N9508 = get_svi_on_device(candidate_vlan_xls_N9508, svi_from_cfg)

print("svi_from_cfg = ", svi_from_cfg)
print("candidate_svi_on_N9508 = ", candidate_svi_on_N9508)

candidate_svi_not_to_be_migrated_N9508 = get_list_not_to_be_migrated(candidate_svi_on_N9508, svi_from_cfg)

print("candidate_svi_not_to_be_migrated_N9508 = ", candidate_svi_not_to_be_migrated_N9508)


################ REAL VLAN  ##############


if not VLAN_FROM_XLS:
    vlan_on_N9508 = list(set(candidate_vlan_xls_N9508) | (set(candidate_vlan_xls_N3048)))
    vlan_on_N9508.sort(key=natural_keys)

else:
    vlan_on_N9508 = candidate_vlan_xls_N9508

vlan_on_N3048 = candidate_vlan_xls_N3048

print("vlan_on_N9508 = ", vlan_on_N9508)
print("vlan_on_N3048 = ", vlan_on_N3048)

if not VLAN_FROM_XLS:
    vlan_not_to_be_migrated_N9508 = get_list_not_to_be_migrated(vlan_on_N9508, vlan_cfg)
else:
    vlan_not_to_be_migrated_N9508 = candidate_vlan_not_to_be_migrated_N9508

vlan_not_to_be_migrated_N3048 = candidate_vlan_not_to_be_migrated_N3048

print("vlan_not_to_be_migrated_N9508 = ", vlan_not_to_be_migrated_N9508)
print("vlan_not_to_be_migrated_N3048 = ", vlan_not_to_be_migrated_N3048)

################ MAIN IF #############

net_in_area_dict = get_net_in_area()
svi_to_area_dict = get_svi_to_area(net_in_area_dict)

svi_on_N9508 = candidate_svi_on_N9508[:]
svi_on_N9508.sort(key=natural_keys)

print("svi_on_N9508 = ", svi_on_N9508)

svi_not_to_be_migrated_N9508 = get_list_not_to_be_migrated(svi_on_N9508, svi_from_cfg)

migr_dict_N9508 = get_migration_dictionary_N9508()
cfg_intf_N9508 = get_normalized_if_OSWVCEVSW_cfg(if_not_to_be_migrated_N9508, migr_dict_N9508, qos_sp_def_N9508_dict)
cfg_vlan_N9508 = get_normalized_vlan_OSWVCEVSW_cfg(vlan_not_to_be_migrated_N9508)
#cfg_svi_N9508 = get_normalized_svi_OSWVCEVSW_cfg(svi_not_to_be_migrated_N9508, svi_on_N9508)
cfg_svi_N9508 = get_normalized_svi_OSWVCEVSW_cfg(svi_not_to_be_migrated_N9508)
cfg_svi_and_ospf_N9508 = add_ospf_to_svi_cfg(cfg_svi_N9508, svi_on_N9508, svi_to_area_dict)
routes_for_N6500 = get_routes_for_devices(routes, svi_on_N9508)
routes_for_N9508 = transform_routes_for_nexus(routes_for_N6500)

migr_dict_N3048 = get_migration_dictionary_N3048()
cfg_intf_N3048 = get_normalized_if_OSWVCEVSW_cfg(if_not_to_be_migrated_N3048, migr_dict_N3048, qos_sp_def_N3048_dict)
cfg_vlan_N3048 = get_normalized_vlan_OSWVCEVSW_cfg(vlan_not_to_be_migrated_N3048)

cfg_N9508 = cfg_vlan_N9508 + cfg_intf_N9508 + cfg_svi_and_ospf_N9508 + routes_for_N9508 + routes_for_N6500
parse_out_N9508 = c.CiscoConfParse(cfg_N9508)
parse_out_N9508.save_as(OSWVCE_CFG_TXT)

cfg_N3048 = cfg_vlan_N3048 + cfg_intf_N3048
parse_out_N3048 = c.CiscoConfParse(cfg_N3048)
parse_out_N3048.save_as(OSWVSW_CFG_TXT)
print("done write")
