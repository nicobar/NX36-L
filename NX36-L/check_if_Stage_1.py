import ciscoconfparse as c
from openpyxl import load_workbook
import re
SWITCH = 'MIOSW058'

SHEET = SWITCH
BASE = '/mnt/hgfs/VM_shared/VF-2017/NMP/'
SITE = 'MI05/'
BASE_DIR = BASE + SITE + SWITCH + '/Stage_1/'
INPUT_XLS = BASE_DIR + SWITCH + '_DB_MIGRATION.xlsx'
OSW_CFG_TXT = BASE_DIR + SWITCH + '.txt'


def atoi(text):
    ''' from string to int'''

    return int(text) if text.isdigit() else text


def natural_keys(text):
    '''
    alist.sort(key=natural_keys) sorts in human order
    http://nedbatchelder.com/blog/200712/human_sorting.html
    (See Toothy's implementation in the comments)
    '''

    return [atoi(c) for c in re.split('(\d+)', text)]


def get_col(ws, col):
    ''' Take a worksheet, return column "col" as lists '''

    return [str(ws.cell(row=r, column=col).value) for r in range(2, ws.max_row)]


parse = c.CiscoConfParse(OSW_CFG_TXT)

intf_obj_list = parse.find_objects(r'^interface .*Ethernet|^interface Port-channel.*')
if_list_cfg = [intf.text for intf in intf_obj_list]

wb = load_workbook(INPUT_XLS)

#ws = wb.get_sheet_by_name(SHEET)
ws = wb[SHEET]

if_list_xls = get_col(ws, 4)

len_if_list_xls = len(if_list_xls)
len_if_list_cfg = len(if_list_cfg)

if_set_xls = set(if_list_xls)
if_set_cfg = set(if_list_cfg)

print(SWITCH)
if (len_if_list_cfg - len_if_list_xls) > 0:
    print("Vi sono più if in cfg che nel xls e quelle in più sono: ")
    cfg_less_xls_list = list(if_set_cfg - if_set_xls)
    cfg_less_xls_list.sort(key=natural_keys)
    print(cfg_less_xls_list)
    for elem in cfg_less_xls_list:
        print(elem)
elif (len_if_list_cfg - len_if_list_xls) == 0:
    print("If in cfg e nel xls sono uguali")
elif (len_if_list_cfg - len_if_list_xls) < 0:
    print("Vi sono più if nel xls che in cfg e quelle in più sono: ")
    xls_less_cfg_list = list(if_set_xls - if_set_cfg)
    xls_less_cfg_list.sort(key=natural_keys)
    print(xls_less_cfg_list)
    for elem in xls_less_cfg_list:
        print(elem)
