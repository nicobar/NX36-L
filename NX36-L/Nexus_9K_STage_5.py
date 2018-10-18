import openpyxl
import sys
from copy import copy

sys.path.insert(0, 'utils')

VLAN_NUMBER = 4096

from get_site_data import get_site_configs, SITES_CONFIG_FOLDER, exists

def get_excel_sheet(filename, sheet_name):
    wb = openpyxl.load_workbook(filename)
    return wb[sheet_name]

def copy_sheet(new_aid, template, sheet):
    i = 0
    print('.', end='')
    for row in range(1, VLAN_NUMBER):
        for column in range(0, ord('Z') - ord('A') + 1):
            i = i + 1
            if i == 10000:
                print('.', end='')
                i = 0
            col = chr(ord('A') + column)
            cell = "{}{}".format(col, row)
            if col == 'Y' and row > 4 and sheet == 'VLAN_Migration_Matrix':
                new_aid[cell].value = ''
                continue
            new_aid[cell].value = template[cell].value
            if template[cell].has_style:
                new_aid[cell].font = copy(template[cell].font)
                new_aid[cell].border = copy(template[cell].border)
                new_aid[cell].fill = copy(template[cell].fill)
                new_aid[cell].number_format = copy(template[cell].number_format)
                new_aid[cell].protection = copy(template[cell].protection)
                new_aid[cell].alignment = copy(template[cell].alignment)

def run(site_configs):
    import re

    AID_PATH = site_configs[0].base_dir + site_configs[0].site + 'AID/'

    site = site_configs[0].switch
    m = re.search('([A-Z]{2})OSW(\d\d)', site)
    site = m.group(1) + m.group(2)

    INPUT_XLS = AID_PATH + 'AID_to_{}_NMP_PARTIAL.xlsx'.format(site)
    OUTPUT_XLS = AID_PATH + 'AID_to_{}_NMP.xlsx'.format(site)

    aid_file_xls = openpyxl.load_workbook(INPUT_XLS)

    matrix_sheet = 'VLAN_Migration_Matrix'
    stp_sheet = 'Higher STP Cost IF'
    STP_Cost = get_excel_sheet(site_configs[0].base_dir + "Migrazione/AID_to_SITE_NMP_TEMPLATE.xlsx", stp_sheet)
    VLAN_Migration_Matrix = get_excel_sheet(site_configs[0].base_dir + "Migrazione/AID_to_SITE_NMP_TEMPLATE.xlsx", matrix_sheet)
    vce1_cell = "{}{}".format("B", "1")
    vce2_cell = "{}{}".format("B", "2")
    osw1_cell = "{}{}".format("B", "3")
    ows2_cell = "{}{}".format("B", "4")
    po69_1_cell = "{}{}".format("D", "1")
    po69_2_cell = "{}{}".format("D", "2")
    STP_Cost[vce1_cell].value = site_configs[0].vpe_router
    STP_Cost[vce2_cell].value = site_configs[1].vpe_router
    STP_Cost[osw1_cell].value = site_configs[0].switch
    STP_Cost[ows2_cell].value = site_configs[1].switch
    STP_Cost[po69_1_cell].value = site_configs[0].vce_switch + '-' + site_configs[0].switch
    STP_Cost[po69_2_cell].value =  site_configs[1].vce_switch + '-' + site_configs[1].switch

    ws = aid_file_xls.create_sheet(matrix_sheet)
    copy_sheet(ws, VLAN_Migration_Matrix, matrix_sheet)
    ws = aid_file_xls.create_sheet(stp_sheet)
    copy_sheet(ws, STP_Cost, stp_sheet)

    aid_file_xls.save(filename=OUTPUT_XLS)

if __name__ == "__main__":
    site_configs = get_site_configs(SITES_CONFIG_FOLDER)
    print('Start Script')
    run(site_configs)
    print('End Script')